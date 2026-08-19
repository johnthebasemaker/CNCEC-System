"""
backend/api/bulk_import.py — Bulk Excel Import (operator "update from Excel").

Four structured workbooook kinds, each with a DRY-RUN (default) that returns a
full plan and a COMMIT mode that applies it in one transaction + audits:

  inventory      CNCEC_Inventory.xlsx sheet "Inventory" → `inventory` master
                 upsert on SAP_Code. Aggregate columns (Receipt/Consumption/
                 Return/Current Stock) are IGNORED — stock is ledger-derived.
                 Category values are canonicalised (e.g. the workbook's
                 "Surface Shield" → the DB's "Surface Shields") so the MTC
                 hard-block keeps matching `mtc_required_category`.
  ledger         The same workbook's "Receipt Log" / "Consumption Log" /
                 "Return Log" sheets → append-only ledger backfill with a
                 three-tier reconcile per row key (day, SAP, qty, ref):
                 exact multiset match → skip; same (day, SAP, ref) with a
                 different qty → UPDATE (workbook corrections, e.g. a
                 mis-entry zeroed out); otherwise INSERT. Rows whose SAP is
                 missing from `inventory` are rejected (soft-FK protection).
                 DB rows absent from the workbook are only REPORTED — this
                 importer never deletes ledger history.
  sme-equipment  Equipment.xlsx "Data Input" → sme_equipment upsert on
                 (Site_ID, tag, code) + sme_sqm_progress re-seed that
                 PRESERVES Done_SQM (ports legacy sme_bootstrap
                 _clean_equipment: Name-identity tag backfill, short-name →
                 code backfill, non-numeric code skip, per-(tag,code) area
                 aggregation with SQM summing, Location canonicalisation).
  sme-recipes    For_1_SQM.xlsx → sme_recipe upsert on (code, material).
  sme-materials  Materials_DetailsAvailable_Qty.xlsx → sme_inventory_seed
                 upsert on (Material_Code, SAP_Code) — one row per PHYSICAL
                 component, quantities summed across that component's PO
                 lines. A multi-part system lists one Material_Code as four
                 Comp-A/B/C/D drums at variant SAPs (2026-07-30 ruling);
                 keying on the code alone summed all four into one figure.
                 SAP-less placeholder rows left by a cutover (the frozen
                 legacy SQLite seed has no SAP_Code column) are retired once
                 the workbook supplies real SAPs for that code.

Two further planners ride on the same workbook and are driven by
`tools/pg_excel_sync.py` rather than by an HTTP endpoint, exactly as the
Surface-Shield routing is:

  plan_rack_locations  Inventory `Rack/Current Location` → storage_locations +
                       material_locations (the store keeper's shelf lookup)
  plan_asset_units     Consumption Log `Location` → asset_units. A Location is
                       what MAKES the row a reusable asset; without one it is
                       ordinary consumption and no unit exists.

Both are create-if-absent — see "THE WORKBOOK SEEDS, THE APP OWNS" below.

Roles: SME kinds are the Master-Data exact-lock {hod, admin}; `inventory` and
`ledger` are admin-only. HOD site pinning follows sme_master._write_site.

⚠️ SHEET SCOPE. `CNCEC_Inventory.xlsx` grew from 5 to 14 sheets on 2026-08-04
(MASTER EQUIPMENTS, Safety Items, RL CONSUMABLES, RL TOOLS AND TACKLES, BR CC
PU TOOLS AND TACKLES, ELECTRICAL ITEMS, INSTRUMENTS, CUMI MATERIAL RECEVIED,
⚙ VBA Setup Guide). Worksheets are selected BY NAME, so those were already
ignored — but `plan_inventory`'s single-sheet fallback read `worksheets[0]`,
which is only `Inventory` by luck of tab order. `_CONSUMED_SHEETS` now makes
the scope explicit and the fallback refuses to guess inside a multi-sheet
workbook (see `_looks_like_cncec_workbook`).
"""
from __future__ import annotations

import io
import re
from collections import Counter
from datetime import date, datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import delete, func, insert, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from .auth import require_level, require_roles
from .db import get_session
from .services.ledger import _MD, write_audit
from .sme_master import _upsert_progress, _write_site

router = APIRouter(prefix="/import", tags=["Bulk Excel import"])

inventory_t = _MD.tables["inventory"]
receipts_t = _MD.tables["receipts"]
consumption_t = _MD.tables["consumption"]
returns_t = _MD.tables["returns"]
equipment_t = _MD.tables["sme_equipment"]
recipe_t = _MD.tables["sme_recipe"]
norm_t = _MD.tables["sme_manpower_norm"]
norm_role_t = _MD.tables["sme_manpower_norm_role"]
roles_t = _MD.tables["mh_roles"]
seed_t = _MD.tables["sme_inventory_seed"]
# 2026-08-05 — the workbook seeds where things live (see `plan_rack_locations`
# and `plan_asset_units`). Both are CREATE-IF-ABSENT: the app owns a place once
# a human has touched it.
storage_loc_t = _MD.tables["storage_locations"]
material_loc_t = _MD.tables["material_locations"]
asset_unit_t = _MD.tables["asset_units"]
asset_move_t = _MD.tables["asset_movements"]

MAX_XLSX_BYTES = 8 * 1024 * 1024

# Workbook category spellings → the DB's canonical values. "Surface Shields"
# drives the MTC hard-block (app_settings.mtc_required_category) — an
# unnormalised singular would silently disarm the gate for imported rows.
CATEGORY_CANON = {"surface shield": "Surface Shields",
                  "surface shields": "Surface Shields",
                  "r/l cons": "R/L Consumables"}

_LOCATION_CANON = {c.lower(): c for c in ("Brown Field", "TRAIN J", "TRAIN K")}

# Lining-system codes were integers ("1", "2") until the 2026-08 workbooks
# renumbered every one of them to a string ("LSC1", "LSC2").
#
# ⚠️ Both planners below used to run `code = str(int(float(code)))` and treat
# the ValueError as "this is a placeholder, skip it". Against the new workbooks
# that predicate is true of EVERY row: the equipment planner skipped all 292
# rows and reported a *warning*, so `tools/pg_excel_sync.py` completed
# successfully having written nothing. Never re-introduce a numeric cast here —
# the column is Text in `models.py` and both SME engines coerce to string.
#
# The placeholder test is now explicit, which is what the cast was only ever
# approximating.
# `_s` above already strips the cell, folds Excel's float-ified "1.0" back to
# "1", and returns None for ''/nan/N/A — so this only has to name the markers
# that ARE a value but name no system yet.
_PLACEHOLDER_CODE_MARKERS = ("tbc", "tbd", "-", "?")


def _is_placeholder_code(code: str) -> bool:
    """True for a cell that names no system yet (To_Be_Confirmed_LSC, TBC, -)."""
    c = (code or "").strip().lower()
    return (not c) or c in _PLACEHOLDER_CODE_MARKERS or c.startswith("to_be_confirmed")



def _s(v: Any) -> Optional[str]:
    """Cell → stripped string or None ('', 'nan', 'None', 'N/A' → None)."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "nan", "None", "N/A", "n/a", "NA"):
        return None
    if s.endswith(".0") and s[:-2].isdigit():  # Excel float-ified codes
        s = s[:-2]
    return s


def _f(v: Any) -> Optional[float]:
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def _iso(v: Any) -> Optional[str]:
    """Excel date/datetime/text → the ledger's 'YYYY-MM-DD HH:MM:SS' text."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d 00:00:00")
    s = _s(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


def _sheet_rows(data: bytes, want: str | None, header_probe: tuple[str, ...],
                required: bool = True) -> tuple[list[str], list[tuple]]:
    """Load one worksheet and find its header row (workbooks carry a title
    banner above the real header). Returns (headers, data_rows)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(422, "not a readable .xlsx workbook")
    try:
        ws = None
        if want is not None:
            for cand in wb.sheetnames:
                if cand.strip().lower() == want.strip().lower():
                    ws = wb[cand]
                    break
            if ws is None:
                if required:
                    raise HTTPException(422, f"worksheet {want!r} not found "
                                             f"(has: {wb.sheetnames})")
                return [], []
        else:
            ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    probe = {p.lower() for p in header_probe}
    for i, row in enumerate(rows[:5]):
        cells = {str(c).strip().lower() for c in row if c is not None}
        if probe <= cells:
            headers = [str(c).strip() if c is not None else "" for c in row]
            return headers, rows[i + 1:]
    raise HTTPException(422, f"header row not found (need columns {sorted(probe)})")


# The ONLY four worksheets this module reads out of CNCEC_Inventory.xlsx.
# Everything else in the workbook (10 sheets as of 2026-08-04) is reference
# material the operator keeps beside the data — never a sync input.
_CONSUMED_SHEETS = ("Inventory", "Receipt Log", "Consumption Log", "Return Log")


def _sheet_names(data: bytes) -> list[str]:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(422, "not a readable .xlsx workbook")
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _looks_like_cncec_workbook(names: list[str]) -> bool:
    """True when the file is the multi-sheet ledger workbook rather than a
    single-sheet master export — i.e. it carries at least one of the log
    sheets. Used to decide whether a missing `Inventory` tab is a DIFFERENT
    kind of file (fall back) or a RENAMED tab in this one (refuse)."""
    lower = {n.strip().lower() for n in names}
    return any(s.lower() in lower for s in _CONSUMED_SHEETS[1:])


def ignored_sheets(data: bytes) -> list[str]:
    """Worksheets present in the workbook that the sync deliberately skips.
    Reported once per run so the scope is visible rather than assumed."""
    consumed = {s.lower() for s in _CONSUMED_SHEETS}
    return [n for n in _sheet_names(data) if n.strip().lower() not in consumed]


def _is_declined_inventory_col(header: str) -> bool:
    """Inventory-sheet columns we have LOOKED AT and chosen not to import.

    Distinguished from merely-unmapped columns so the warning text can say
    which it is:

      `Current Location`  BLANK in 452 of 452 rows on the 2026-08-04 workbook.
                          The spreadsheet cannot seed locations, so the app
                          owns them — racks in `storage_locations`, per-unit
                          whereabouts in `asset_units`.
      `Audit <date>`      a dated physical-count column the operator adds per
                          stock take. Stock is ledger-derived; importing a
                          count as master data would create a second truth
                          (the workbook already disagrees with the DB on 131
                          of 452 SAPs).
    """
    h = header.strip().lower()
    return h == "current location" or h.startswith("audit ")


def _col(headers: list[str], *names: str) -> Optional[int]:
    lower = [h.lower() for h in headers]
    for n in names:
        if n.lower() in lower:
            return lower.index(n.lower())
    return None


# ─── inventory master ─────────────────────────────────────────────────────────
async def plan_inventory(session: AsyncSession, data: bytes, site_id: str) -> dict:
    headers, rows = _sheet_rows(data, "Inventory",
                                ("sap code", "category", "opening stock"),
                                required=False)
    if not headers:
        # Single-sheet master-file fallback — but ONLY for a file that is not
        # the CNCEC ledger workbook. Inside that workbook a missing `Inventory`
        # tab means the tab was renamed or reordered, and `worksheets[0]` would
        # then be `Request` or `Safety Items`; every write is
        # `ON CONFLICT DO UPDATE`, so the master would be quietly rebuilt from
        # the wrong sheet. Fail loudly with the sheet list instead.
        names = _sheet_names(data)
        if _looks_like_cncec_workbook(names):
            raise HTTPException(
                422, "worksheet 'Inventory' not found in what looks like the "
                     f"CNCEC ledger workbook (has: {names}). Refusing to guess "
                     "a sheet — rename the tab back to 'Inventory'.")
        headers, rows = _sheet_rows(data, None, ("sap code", "category"))
    colspec = {
        "sl": ("Sl. No.", "Sl_No", "Sl. #"), "sap": ("SAP CODE", "SAP_Code"),
        "mat": ("Material Code", "Material_Code"),
        "desc": ("Equipment Description", "Equipment_Description"),
        "uom": ("UOM",), "cat": ("Category",),
        "open": ("Opening Stock", "Opening_Stock"),
        "min": ("Minimum Qty", "Minimum_Qty"),
    }
    ix = {k: _col(headers, *names) for k, names in colspec.items()}
    if ix["sap"] is None:
        raise HTTPException(422, "SAP CODE column missing")
    # Columns are resolved by NAME (order-independent). Anything the workbook
    # carries beyond the mapped set + the ledger-derived aggregates is ignored
    # — but LOUDLY, so a restructured sheet never loses data silently.
    _known = {n.lower() for names in colspec.values() for n in names}
    _known |= {"receipt", "consumption", "return", "current stock"}
    # Consumed by `plan_rack_locations`, which writes storage_locations rather
    # than a column on `inventory` — a material legitimately sits in more than
    # one rack, which is the wrong grain for this one-row-per-SAP table.
    _known |= {n.lower() for n in _RACK_COL_NAMES}
    # Two classes of unmapped column, reported DIFFERENTLY on purpose. Both are
    # still reported — a restructured sheet must never drop data silently — but
    # a column we have analysed and rejected is not the same finding as one
    # nobody has looked at, and collapsing them trains the operator to ignore
    # the warning that matters.
    extra_cols, declined_cols = [], []
    for h in headers:
        if not h or h.lower() in _known:
            continue
        (declined_cols if _is_declined_inventory_col(h) else extra_cols).append(h)

    existing = {r["SAP_Code"]: dict(r) for r in
                (await session.execute(select(inventory_t))).mappings().all()}
    mat_owner = {r["Material_Code"]: r["SAP_Code"] for r in existing.values()
                 if r.get("Material_Code")}

    # First pass — collect every row so Material_Code conflicts can be judged
    # against the WHOLE file (a code may legitimately change owners when its
    # current owner is re-mapped in the same workbook).
    parsed: list[dict] = []
    rejects, warnings = [], []
    if extra_cols:
        warnings.append("ignored unmapped column(s): " + ", ".join(extra_cols))
    if declined_cols:
        warnings.append("column(s) deliberately NOT imported: "
                        + ", ".join(declined_cols)
                        + " — see bulk_import._is_declined_inventory_col")
    normalised_cats = Counter()
    seen_saps = set()
    for n, row in enumerate(rows, start=1):
        sap = _s(row[ix["sap"]]) if ix["sap"] < len(row) else None
        if not sap:
            continue
        if sap in seen_saps:
            rejects.append({"row": n, "sap": sap, "reason": "duplicate SAP in file"})
            continue
        seen_saps.add(sap)

        def cell(key):
            i = ix[key]
            return row[i] if i is not None and i < len(row) else None

        cat_raw = _s(cell("cat"))
        cat = CATEGORY_CANON.get(cat_raw.lower(), cat_raw) if cat_raw else None
        if cat_raw and cat != cat_raw:
            normalised_cats[f"{cat_raw} → {cat}"] += 1
        parsed.append({"row": n, "sap": sap, "mat": _s(cell("mat")),
                       "fields": {"Sl_No": _s(cell("sl")),
                                  "Equipment_Description": _s(cell("desc")),
                                  "UOM": _s(cell("uom")), "Category": cat,
                                  "Opening_Stock": _f(cell("open")),
                                  "Minimum_Qty": _f(cell("min"))}})

    # Material_Code resolution (unique across inventory):
    #   in-file duplicate        → first row keeps it, later rows import codeless
    #   owner re-mapped in file  → release + reassign (the workbook is truth)
    #   owner NOT in the file    → keep the owner's code; import row codeless
    file_mat: dict[str, str] = {}
    for p in parsed:
        if p["mat"] and p["mat"] not in file_mat:
            file_mat[p["mat"]] = p["sap"]
    file_saps = {p["sap"] for p in parsed}
    file_mat_of_sap = {p["sap"]: p["mat"] for p in parsed}
    releases: list[dict] = []
    for p in parsed:
        mat = p["mat"]
        if not mat:
            continue
        if file_mat.get(mat) != p["sap"]:
            warnings.append(f"SAP {p['sap']}: Material_Code {mat} already used by "
                            f"SAP {file_mat[mat]} in this file — imported without it")
            p["mat"] = None
            continue
        owner = mat_owner.get(mat)
        if owner and owner != p["sap"]:
            if owner in file_saps and file_mat_of_sap.get(owner) != mat:
                releases.append({"sap": owner, "mat": mat})  # re-mapped → release
            else:
                warnings.append(f"SAP {p['sap']}: Material_Code {mat} stays with "
                                f"SAP {owner} (not re-mapped here) — imported "
                                f"without it")
                p["mat"] = None

    inserts, updates, unchanged = [], [], 0
    for p in parsed:
        fields = {k: v for k, v in p["fields"].items() if v is not None}
        if p["mat"] is not None:
            fields["Material_Code"] = p["mat"]
        cur = existing.get(p["sap"])
        if cur is None:
            inserts.append({"SAP_Code": p["sap"], "Site_ID": site_id, **fields})
        else:
            diff = {k: v for k, v in fields.items() if cur.get(k) != v}
            if diff:
                if "Opening_Stock" in diff:
                    warnings.append(f"SAP {p['sap']}: Opening_Stock "
                                    f"{cur.get('Opening_Stock')} → {diff['Opening_Stock']}")
                updates.append({"SAP_Code": p["sap"], "diff": diff})
            else:
                unchanged += 1
    if normalised_cats:
        warnings.append("category canonicalised: " +
                        ", ".join(f"{k} ×{v}" for k, v in normalised_cats.items()))
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": rejects, "warnings": warnings, "releases": releases}


async def apply_inventory(session: AsyncSession, plan: dict, username: str) -> None:
    # Free re-mapped Material_Codes FIRST so the unique constraint never trips
    # mid-plan (their new values arrive with the updates below).
    for rel in plan.get("releases", []):
        await session.execute(update(inventory_t)
                              .where(inventory_t.c["SAP_Code"] == rel["sap"],
                                     inventory_t.c["Material_Code"] == rel["mat"])
                              .values(Material_Code=None))
    for row in plan["inserts"]:
        await session.execute(insert(inventory_t).values(**row))
    for u in plan["updates"]:
        await session.execute(update(inventory_t)
                              .where(inventory_t.c["SAP_Code"] == u["SAP_Code"])
                              .values(**u["diff"]))
    await write_audit(session, username, "BULK_IMPORT_INVENTORY", "inventory",
                      f"+{len(plan['inserts'])} ~{len(plan['updates'])} "
                      f"rejected={len(plan['rejects'])}")


# ─── ledger backfill (receipts / consumption / returns) ───────────────────────
# Every sheet's columns resolve by NAME (order-independent). "ignore" lists
# workbook columns that deliberately have no DB home — anything else that is
# unmapped raises a warning so a restructured sheet never drops data silently.
# ("Material Code" / "Equipment Description" / "UOM" repeat the inventory
# master on every log sheet and are always ignored.)
_LEDGER_ALWAYS_IGNORED = ("date", "sap code", "sap_code", "material code",
                          "equipment description", "uom")
_LEDGER_SHEETS = {
    "receipts": {
        "sheet": "Receipt Log", "table": receipts_t,
        "cols": {"Quantity": ("Qty.",), "Serial_No": ("Serial No.",),
                 "PR_Number": ("PR#",), "WBS": ("WBS#",),
                 "Location": ("Location",), "Vehicle_No": ("Vehicle No.",),
                 "Driver_Name": ("Driver Name",), "DN_No": ("DN. No.",),
                 "Pallet_No": ("Pallet No.",), "Mob_From": ("Mob. From",),
                 "Mob_To": ("Mob. To",), "Prepared_by": ("Prepared by",),
                 "Received_by": ("Received by",), "DN_Copy": ("DN. Copy",),
                 "Remarks": ("Remarks",)},
        "ref": "DN_No", "ignore": (),
    },
    "consumption": {
        "sheet": "Consumption Log", "table": consumption_t,
        "cols": {"Quantity": ("Qty.",), "Serial_No": ("Serial No.",),
                 "PR_Number": ("PR#",), "Work_Type": ("Work Type",),
                 "Tank_No": ("Tank No.",), "WBS": ("WBS#",),
                 "Approved By": ("Approved By",), "Issued_To": ("Received by",),
                 "Issued_By": ("Prepared by",), "Remarks": ("Remarks",),
                 # 2026-08-04: the programme a consumption belongs to. All
                 # 1,110 rows carry it, and it is what routes Surface Shields
                 # into the SME portal — see `plan_sme_routing`.
                 "Item_Type": ("type",)},
        # `consumption` has no Pallet_No / paper-number columns — 2026-07-14
        # workbook restructure adds both to the sheet; ignored by design.
        #
        # 2026-08-04 additions, both deliberately unmapped ONTO `consumption`:
        #   `Location`      2026-08-05 ruling: a Location makes the row a
        #                   REUSABLE ASSET, so it is read by
        #                   `plan_asset_units` and written to `asset_units` —
        #                   never onto the ledger row, which records an event
        #                   and not a whereabouts.
        #   `Current Stock` a spreadsheet formula result. `Opening_Stock +
        #                   Σledger` is already computed server-side, and
        #                   importing this would create a second, divergent
        #                   truth (the workbook already disagrees with the DB
        #                   on 131 of 452 SAPs).
        "ref": "Tank_No", "ignore": ("cons. paper no.", "pallet no.",
                                     "location", "current stock"),
    },
    "returns": {
        "sheet": "Return Log", "table": returns_t,
        "cols": {"Quantity": ("Qty.",), "Reason": ("Reason",),
                 "Remarks": ("Remarks",)},
        # the Return Log reuses the Receipt Log template; `returns` is a
        # narrow table (Date/SAP/Qty/Reason/Remarks) so the rest has no home
        "ref": "Reason",
        "ignore": ("serial no.", "pr#", "wbs#", "location", "vehicle no.",
                   "driver name", "dn. no.", "pallet no.", "mob. from",
                   "mob. to", "prepared by", "received by", "dn. copy"),
    },
}


def _day(v) -> str:
    return str(v or "")[:10]


async def plan_ledger(session: AsyncSession, data: bytes, site_id: str,
                      extra_saps: set[str] | None = None) -> dict:
    known_saps = {r[0] for r in
                  (await session.execute(select(inventory_t.c["SAP_Code"]))).all()}
    known_saps |= extra_saps or set()  # dry-run chained after an inventory plan
    out = {"sections": {}, "rejects": [], "warnings": []}
    for kind, spec in _LEDGER_SHEETS.items():
        headers, rows = _sheet_rows(data, spec["sheet"], ("sap code", "qty."),
                                    required=False)
        section = {"inserts": [], "corrections": [], "matched": 0,
                   "zero_skipped": 0, "db_only": 0}
        out["sections"][kind] = section
        if not headers:
            out["warnings"].append(f"{spec['sheet']}: sheet missing — skipped")
            continue
        sap_i = _col(headers, "SAP CODE", "SAP_Code")
        date_i = _col(headers, "Date", "Date ")
        colmap = {field: _col(headers, *names)
                  for field, names in spec["cols"].items()}
        _known = set(_LEDGER_ALWAYS_IGNORED) | set(spec["ignore"])
        _known |= {n.lower() for names in spec["cols"].values() for n in names}
        extra_cols = [h for h in headers if h and h.lower() not in _known]
        if extra_cols:
            out["warnings"].append(f"{spec['sheet']}: ignored unmapped "
                                   f"column(s): {', '.join(extra_cols)}")

        file_rows = []
        for n, row in enumerate(rows, start=1):
            sap = _s(row[sap_i]) if sap_i is not None and sap_i < len(row) else None
            if not sap:
                continue
            d = _iso(row[date_i]) if date_i is not None and date_i < len(row) else None
            qty = _f(row[colmap["Quantity"]]) if colmap["Quantity"] is not None else None
            if d is None or qty is None:
                out["rejects"].append({"sheet": spec["sheet"], "row": n, "sap": sap,
                                       "reason": "missing/unparseable Date or Qty"})
                continue
            if sap not in known_saps:
                out["rejects"].append({"sheet": spec["sheet"], "row": n, "sap": sap,
                                       "reason": "SAP not in inventory master"})
                continue
            vals = {"Date": d, "SAP_Code": sap, "Quantity": qty, "Site_ID": site_id}
            for field, i in colmap.items():
                if field == "Quantity" or i is None or i >= len(row):
                    continue
                v = _s(row[i])
                if v is not None:
                    vals[field] = v
            file_rows.append(vals)

        table, ref = spec["table"], spec["ref"]
        db_rows = [dict(m) for m in (await session.execute(
            select(table).where(table.c["Site_ID"] == site_id))).mappings().all()]

        def key(r):  # exact multiset identity
            return (_day(r.get("Date")), r.get("SAP_Code"),
                    round(float(r.get("Quantity") or 0), 4),
                    _s(r.get(ref)) or "")

        def refkey(r):  # correction identity: same day+sap+ref, any qty
            return (_day(r.get("Date")), r.get("SAP_Code"), _s(r.get(ref)) or "")

        db_exact = Counter(key(r) for r in db_rows)
        # tier 1 — exact matches consume DB copies
        remaining = []
        for fr in file_rows:
            k = key(fr)
            if db_exact.get(k, 0) > 0:
                db_exact[k] -= 1
                section["matched"] += 1
            else:
                remaining.append(fr)
        # unmatched DB copies, grouped for the correction tier
        db_left: dict[tuple, list[dict]] = {}
        for r in db_rows:
            k = key(r)
            if db_exact.get(k, 0) > 0:
                db_exact[k] -= 1
                db_left.setdefault(refkey(r), []).append(r)
        # tier 2 — qty corrections (workbook is truth for the same day+sap+ref)
        inserts = []
        for fr in remaining:
            cands = db_left.get(refkey(fr)) or []
            if cands:
                target = cands.pop(0)
                section["corrections"].append(
                    {"id": target["id"], "sap": fr["SAP_Code"],
                     "date": _day(fr["Date"]),
                     "qty_from": target["Quantity"], "qty_to": fr["Quantity"]})
            elif float(fr["Quantity"]) == 0.0:
                section["zero_skipped"] += 1  # zero-qty history line, no DB twin
            else:
                inserts.append(fr)
        section["inserts"] = inserts
        section["db_only"] = sum(len(v) for v in db_left.values())
        if section["db_only"]:
            out["warnings"].append(
                f"{spec['sheet']}: {section['db_only']} DB row(s) have no workbook "
                f"counterpart — left untouched (this importer never deletes)")
    return out


# ─── Surface-Shield routing → the SME portal (rule 1a safe) ───────────────────
#
# WHAT THIS DOES AND, MORE IMPORTANTLY, WHAT IT DOES NOT.
#
# Consumption rows tagged `type = 'Surface Shield'` are the SME programme's
# actual physical draw. They are recorded in `sme_consumption_log` — a
# REPORTING table read only by the variance comparison (sme.SQL_SME_COMPARISON)
# and a Man-Hours rollup.
#
# ⚠️ **`sme_inventory_seed` IS NEVER WRITTEN HERE.** Rule 1a (SME ⇄ ERP strict
# decoupling) says every SME quantity comes from that seed and nowhere else, so
# a warehouse issue must not move a single estimator figure. Logging actual
# consumption alongside the plan is not the same thing as netting it off the
# plan, and this module does the first and never the second. The UI shows the
# logged draw as a SIDE NOTE ("Actual Physical Balance") beside the estimator's
# numbers — deliberately adjacent, never merged.
#
# Suite BA greps `SQL_SME_MATERIALS` + `_CALC_POOL_SQL` for ERP table names and
# now also for `sme_consumption_log`, so a future join is caught at review.

_SURFACE_SHIELD_TYPE = "surface shield"

# Aliases an operator should never be asked to map to equipment — they are
# places or activities, not vessels. Pre-marked `ignored` so the resolve screen
# shows only the real questions.
_NON_EQUIPMENT_ALIASES = {
    "TOSITE", "HOUSEKEEPING", "INYARD", "SCAFFOLDING", "OTHERS", "OTHER",
    "SAMPLEPLATE", "NA", "SITE",
}


def alias_norm(v) -> str:
    """`Tank No.` → its matching form.

    Upper-cases, drops every separator, then strips leading zeros inside each
    run of digits. That last step is what collapses the four spellings of one
    tank — `J091` / `J0091` / `J-0091` / `J 0091` all become `J91` — which is
    58 of the 103 Surface-Shield rows. Without it they read as four vessels.
    """
    s = re.sub(r"[^A-Za-z0-9]", "", str(v or "")).upper()
    return re.sub(r"0*(\d+)", lambda m: m.group(1), s)


def match_alias(norm: str, tag_index: dict[str, set[str]]) -> list[str]:
    """Equipment tags an alias could mean, best interpretation first.

    EXACT normalised match wins outright. Only when there is none do we try a
    suffix match, and a suffix match that hits more than one tag is reported as
    ambiguous rather than resolved — `TNK-091` suffix-matches BOTH
    `522-8J10-TNK-091` (TRAIN J) and `522-8k10-TNK-091` (TRAIN K), and those
    are two different vessels on two different trains. Picking one would put 39
    rows of real consumption against the wrong train and look entirely
    plausible in every report afterwards.
    """
    if not norm or norm in _NON_EQUIPMENT_ALIASES:
        return []
    exact = tag_index.get(norm)
    if exact:
        return sorted(exact)
    return sorted({t for k, tags in tag_index.items()
                   if k.endswith(norm) and k != norm for t in tags})


async def plan_sme_routing(session: AsyncSession, site_id: str,
                           ledger_plan: dict) -> dict:
    """Plan the Surface-Shield half of a ledger sync.

    Derived from the consumption rows the ledger plan is ABOUT TO INSERT, which
    is what makes it idempotent for free: `plan_ledger`'s three-tier reconcile
    already skips rows that exist, so a re-run inserts nothing and therefore
    logs nothing.
    """
    out = {"aliases": [], "log_inserts": [], "warnings": [],
           "skipped_not_sme": 0, "unassigned": 0}
    rows = [r for r in (ledger_plan.get("sections", {})
                        .get("consumption", {}).get("inserts") or [])
            if str(r.get("Item_Type") or "").strip().lower() == _SURFACE_SHIELD_TYPE]
    if not rows:
        return out

    alias_t = _MD.tables["sme_tank_alias"]

    tag_index: dict[str, set[str]] = {}
    for (tag,) in (await session.execute(
            select(equipment_t.c["Equipment_Tag_No"]).distinct()
            .where(equipment_t.c["Site_ID"] == site_id))).all():
        tag_index.setdefault(alias_norm(tag), set()).add(tag)

    # SAP → Material_Code, then the seed's codes: only a material the estimator
    # actually models can carry an Expected_Qty.
    sap_to_mat = {r[0]: r[1] for r in (await session.execute(
        select(inventory_t.c["SAP_Code"], inventory_t.c["Material_Code"]))).all()
        if r[1]}
    seed_codes = {r[0] for r in (await session.execute(
        select(seed_t.c["Material_Code"]).distinct())).all()}

    existing = {r["alias_norm"]: dict(r) for r in (await session.execute(
        select(alias_t).where(alias_t.c["Site_ID"] == site_id))).mappings().all()}

    # ── alias registry ──
    counts = Counter(_s(r.get("Tank_No")) or "" for r in rows)
    for raw, n in sorted(counts.items()):
        norm = alias_norm(raw)
        if not norm:
            continue
        cands = match_alias(norm, tag_index)
        prior = existing.get(norm)
        if prior and prior.get("status") in ("mapped", "ignored"):
            continue          # an operator's decision is never overwritten
        entry = {"Site_ID": site_id, "alias_raw": raw, "alias_norm": norm,
                 "match_count": len(cands), "row_count": n}
        if norm in _NON_EQUIPMENT_ALIASES:
            entry |= {"status": "ignored", "Equipment_Tag_No": None}
        elif len(cands) == 1:
            entry |= {"status": "mapped", "Equipment_Tag_No": cands[0]}
        else:
            entry |= {"status": "unresolved", "Equipment_Tag_No": None}
            out["warnings"].append(
                f"tank alias {raw!r}: {n} row(s) held — "
                + (f"matches {len(cands)} equipment tags ({', '.join(cands)})"
                   if cands else "no equipment tag matches")
                + " — resolve in SME → Tank Aliases")
        out["aliases"].append(entry)

    resolved = {a["alias_norm"]: a.get("Equipment_Tag_No")
                for a in out["aliases"] if a["status"] == "mapped"}
    for norm, prior in existing.items():
        if prior.get("status") == "mapped" and prior.get("Equipment_Tag_No"):
            resolved[norm] = prior["Equipment_Tag_No"]

    # ── the log rows ──
    batch = f"xlsx-{site_id}-{datetime.now():%Y%m%dT%H%M%S}"
    for r in rows:
        mat = sap_to_mat.get(r.get("SAP_Code"))
        if not mat or mat not in seed_codes:
            # A Surface-Shield consumable outside the estimator's recipe set —
            # it belongs in the ERP ledger only. Counted, never silently lost.
            out["skipped_not_sme"] += 1
            continue
        tag = resolved.get(alias_norm(r.get("Tank_No"))) or ""
        if not tag:
            out["unassigned"] += 1
        out["log_inserts"].append({
            "batch_id": batch, "Site_ID": site_id,
            "entry_date": _day(r.get("Date")), "entered_by": "excel-sync",
            # '' — not NULL: both columns are NOT NULL on this table, and ''
            # is the codebase's existing "not scoped yet" sentinel. The
            # `unassigned` status is what the UI filters on.
            "Equipment_Tag_No": tag, "Lining_System_Code": "",
            "Material_Code": mat,
            # The workbook states a QUANTITY issued, never an area covered, so
            # SQM_Completed stays 0 until an operator records it on the
            # assignment screen. Expected_Qty needs a system code, so it stays
            # 0 too — a variance against a guessed system is worse than none.
            "SQM_Completed": 0.0, "Expected_Qty": 0.0,
            "Actual_Qty": float(r.get("Quantity") or 0),
            "status": "committed" if tag else "unassigned",
            "notes": f"Tank No. {_s(r.get('Tank_No')) or '—'} · SAP "
                     f"{r.get('SAP_Code')} · from {r.get('Date')}",
        })
    if out["skipped_not_sme"]:
        out["warnings"].append(
            f"{out['skipped_not_sme']} Surface-Shield row(s) are not estimator "
            "materials (no sme_inventory_seed entry) — ERP ledger only")
    if out["unassigned"]:
        out["warnings"].append(
            f"{out['unassigned']} Surface-Shield row(s) logged UNASSIGNED — "
            "assign equipment + SQM in SME → Actual Consumption")
    return out


async def apply_sme_routing(session: AsyncSession, plan: dict,
                            username: str) -> None:
    """Write the alias registry and the SME consumption log.

    Aliases upsert on (Site_ID, alias_norm) and never clobber a row an operator
    has already decided — `plan_sme_routing` filters those out, and the
    `DO UPDATE` here only refreshes the counts.
    """
    alias_t = _MD.tables["sme_tank_alias"]
    log_t = _MD.tables["sme_consumption_log"]
    for a in plan.get("aliases", []):
        stmt = pg_insert(alias_t).values(**a)
        await session.execute(stmt.on_conflict_do_update(
            constraint="uq_sme_tank_alias_site_norm",
            set_={"alias_raw": stmt.excluded.alias_raw,
                  "match_count": stmt.excluded.match_count,
                  "row_count": stmt.excluded.row_count,
                  "status": stmt.excluded.status,
                  "Equipment_Tag_No": stmt.excluded["Equipment_Tag_No"]}))
    for row in plan.get("log_inserts", []):
        await session.execute(insert(log_t).values(**row))
    if plan.get("aliases") or plan.get("log_inserts"):
        await write_audit(
            session, username, "SME_ROUTING", "sme_consumption_log",
            f"+{len(plan.get('log_inserts', []))} Surface-Shield log row(s), "
            f"{len(plan.get('aliases', []))} tank alias/es "
            f"({plan.get('unassigned', 0)} unassigned)")


# ─── where things live: racks (Inventory) and assets (Consumption Log) ────────
#
# Two planners, one ruling: **THE WORKBOOK SEEDS, THE APP OWNS.**
#
# A spreadsheet cell is a starting point typed by whoever last edited the file.
# A row in `asset_units` or `material_locations` is where a store keeper says a
# thing actually is — often after walking to it and scanning its label, with a
# GPS fix attached. Those are not the same claim, and the second one wins. So
# both planners CREATE what is missing and never overwrite what is there:
#
#   storage_locations   ON CONFLICT (Site_ID, code) DO NOTHING — an operator's
#                       zone/rack/row breakdown survives every re-sync.
#   material_locations  a SAP that already has ANY assignment is left alone.
#   asset_units         an existing unit keeps its status, its rack and above
#                       all its `current_lat`/`current_lng`. The workbook's
#                       Location text refreshes `location_note` only while the
#                       app has never touched the unit (`last_seen_by` is still
#                       the sync itself and no fix has been recorded).
#
# ⚠️ STATE OF THE WORKBOOK, 2026-08-05. Both columns exist and both are
# effectively empty: `Rack/Current Location` is blank in 453 of 453 Inventory
# rows, and `Location` is filled on 1 of 1,166 Consumption Log rows ("At site",
# on a row with no Serial No., so it cannot be keyed and is reported back). A
# run today therefore seeds nothing — that is the correct outcome, not a
# failure, and both planners are written to be no-ops on a blank column and on
# a missing one alike.

# The workbook writes a condition in prose. `asset_units.status` is a free-text
# column, so these map onto the vocabulary the app's own picker offers rather
# than onto the custody values — an operator saying "not in use" means the
# hammer is idle, not that it was returned to stores.
_ASSET_STATUS_CANON = {
    "working": "working", "in use": "working", "in-use": "working",
    "ok": "working", "good": "working", "active": "working",
    "not in use": "not_in_use", "not-in-use": "not_in_use",
    "notinuse": "not_in_use", "idle": "not_in_use", "unused": "not_in_use",
    "spare": "not_in_use",
    "repair": "repair", "under repair": "repair", "maintenance": "repair",
    "service": "repair", "damaged": "repair",
    "lost": "lost", "missing": "lost",
    "scrapped": "scrapped", "scrap": "scrapped", "condemned": "scrapped",
    "in_stock": "in_stock", "in stock": "in_stock", "stock": "in_stock",
    "issued": "issued", "returned": "returned",
}

_RACK_COL_NAMES = ("Rack/Current Location", "Rack / Current Location",
                   "Rack/Location", "Rack No", "Rack_No", "Rack")


async def plan_rack_locations(session: AsyncSession, data: bytes, site_id: str,
                              extra_saps: set[str] | None = None) -> dict:
    """Inventory sheet → `storage_locations` + `material_locations`.

    The workbook states one free-text place per SAP. That text becomes the
    rack's `code` (the QR payload printed on the shelf) AND its `description`,
    and the zone / rack / row / bin breakdown is left EMPTY on purpose: one
    column cannot be split into four without guessing, and `_label()` already
    falls back to the description, so an unparsed rack still reads correctly on
    the locator. An operator who wants the breakdown fills it in the app, and
    `DO NOTHING` means the next sync will not undo that.

    Returns a plan even when the column is absent — a workbook without it is a
    valid workbook, not an error.
    """
    out = {"racks": [], "links": [], "warnings": [], "kept": 0, "blank": 0,
           "rejects": [], "column": None}
    headers, rows = _sheet_rows(data, "Inventory",
                                ("sap code", "category"), required=False)
    if not headers:
        out["warnings"].append("Inventory sheet missing — no racks seeded")
        return out
    rack_i = _col(headers, *_RACK_COL_NAMES)
    if rack_i is None:
        out["warnings"].append(
            f"no rack column (looked for {' / '.join(_RACK_COL_NAMES)}) "
            f"— no racks seeded")
        return out
    out["column"] = headers[rack_i]
    sap_i = _col(headers, "SAP CODE", "SAP_Code")
    if sap_i is None:
        raise HTTPException(422, "SAP CODE column missing")

    # `extra_saps` are the SAPs the inventory plan is about to insert. Without
    # them a brand-new material would have its rack rejected on the very run
    # that introduces it, and only pick one up on the NEXT sync.
    known_saps = {r[0].strip() for r in (await session.execute(
        select(inventory_t.c["SAP_Code"]))).all() if r[0]}
    known_saps |= extra_saps or set()
    existing_codes = {r[0] for r in (await session.execute(
        select(storage_loc_t.c["code"])
        .where(storage_loc_t.c["Site_ID"] == site_id))).all()}
    assigned = {r[0] for r in (await session.execute(
        select(material_loc_t.c["SAP_Code"]).distinct()
        .where(material_loc_t.c["Site_ID"] == site_id))).all()}

    seen_codes: set[str] = set()
    seen_saps: set[str] = set()
    for n, row in enumerate(rows, start=1):
        sap = _s(row[sap_i]) if sap_i < len(row) else None
        place = _s(row[rack_i]) if rack_i < len(row) else None
        if not sap:
            continue
        if not place:
            out["blank"] += 1
            continue
        code = re.sub(r"\s+", " ", place).strip()
        if sap not in known_saps:
            out["rejects"].append({"row": n, "sap": sap,
                                   "reason": "SAP not in inventory master"})
            continue
        if code not in existing_codes and code not in seen_codes:
            seen_codes.add(code)
            out["racks"].append({"Site_ID": site_id, "code": code,
                                 "description": code, "created_by": "excel-sync"})
        if sap in assigned:
            # Somebody has already said where this lives, in the app. The
            # workbook does not get to move it.
            out["kept"] += 1
            continue
        if sap in seen_saps:
            continue
        seen_saps.add(sap)
        out["links"].append({"Site_ID": site_id, "SAP_Code": sap, "code": code,
                             "is_primary": True, "note": "seeded from workbook",
                             "updated_by": "excel-sync"})
    if out["kept"]:
        out["warnings"].append(
            f"{out['kept']} SAP(s) already have a rack assigned in the app — "
            f"the workbook's value was NOT applied (the app owns a place once "
            f"a human has set it)")
    return out


async def apply_rack_locations(session: AsyncSession, plan: dict,
                               username: str) -> dict:
    """Create the missing racks, then link the materials to them.

    Racks first and flushed, because the links resolve `code` → `id` against
    rows this same call has just inserted.
    """
    counts = {"racks": 0, "links": 0}
    for rack in plan.get("racks", []):
        stmt = pg_insert(storage_loc_t).values(**rack)
        await session.execute(stmt.on_conflict_do_nothing(
            constraint="uq_storage_locations_site_code"))
        counts["racks"] += 1
    if plan.get("racks"):
        await session.flush()

    if plan.get("links"):
        ids = {r[0]: r[1] for r in (await session.execute(
            select(storage_loc_t.c["code"], storage_loc_t.c["id"])
            .where(storage_loc_t.c["Site_ID"] == plan["links"][0]["Site_ID"]))).all()}
        for link in plan["links"]:
            loc_id = ids.get(link["code"])
            if loc_id is None:          # the rack row vanished under us
                continue
            vals = {k: v for k, v in link.items() if k != "code"}
            stmt = pg_insert(material_loc_t).values(location_id=loc_id, **vals)
            await session.execute(stmt.on_conflict_do_nothing(
                constraint="uq_material_locations_site_sap_loc"))
            counts["links"] += 1
    if counts["racks"] or counts["links"]:
        await write_audit(session, username, "BULK_IMPORT_RACKS",
                          "material_locations",
                          f"pg_excel_sync: +{counts['racks']} rack(s), "
                          f"+{counts['links']} material link(s)")
    return counts


async def plan_asset_units(session: AsyncSession, data: bytes, site_id: str,
                           extra_saps: set[str] | None = None) -> dict:
    """Consumption Log → `asset_units`, one row per physical thing.

    ═══════════════════════════════════════════════════════════════════════════
    THE GOLDEN RULE: a `Location` makes the row an ASSET. No Location, and it
    is ordinary consumption — no unit is created, and the ledger keeps it.
    ═══════════════════════════════════════════════════════════════════════════

    That single test is what separates a hammer from a drum of primer without
    needing a second column to say which is which, and it is the operator's own
    convention rather than one this code invented.

    KEYED ON `(Site_ID, SAP_Code, Serial No.)` — the constraint that already
    exists on the table. Two hammers share a SAP and are told apart by serial;
    that is the whole reason the table is not just a column on `inventory`.

    ⚠️ A row with a Location but NO serial cannot be keyed and is NOT invented
    a serial for: two such rows would silently become one asset, or one asset
    would be created twice on the next run. They are counted and named back to
    the operator instead, which is a fixable spreadsheet problem rather than a
    permanent data one. (The 2026-08-05 workbook has exactly one such row.)

    This planner reads the SHEET, not the ledger plan — unlike
    `plan_sme_routing`, which derives from the rows about to be inserted. An
    asset is STATE, not an event: it must be seedable from a workbook whose
    consumption rows are already loaded, and re-running must converge. The
    create-if-absent write is what makes that idempotent.
    """
    out = {"inserts": [], "seen_notes": [], "warnings": [], "rejects": [],
           "no_serial": [], "consumable_rows": 0, "existing": 0,
           "duplicate_rows": 0, "columns": {}}
    headers, rows = _sheet_rows(data, "Consumption Log", ("sap code", "qty."),
                                required=False)
    if not headers:
        out["warnings"].append("Consumption Log sheet missing — no assets seeded")
        return out
    loc_i = _col(headers, "Location", "Current Location")
    if loc_i is None:
        out["warnings"].append(
            "Consumption Log has no Location column — nothing marks a row as a "
            "reusable asset, so none were seeded")
        return out
    sap_i = _col(headers, "SAP CODE", "SAP_Code")
    ser_i = _col(headers, "Serial No.", "Serial_No", "Serial No")
    st_i = _col(headers, "Status", "Condition")
    date_i = _col(headers, "Date")
    out["columns"] = {"location": headers[loc_i],
                      "serial": headers[ser_i] if ser_i is not None else None,
                      "status": headers[st_i] if st_i is not None else None}
    if ser_i is None:
        out["warnings"].append(
            "Consumption Log has no Serial No. column — an asset cannot be told "
            "apart from another of the same SAP, so none were seeded")
        return out
    if st_i is None:
        # Expected today: the operator sets condition in the app instead.
        out["warnings"].append(
            "no Status column — seeded units start 'in_stock'; set the real "
            "condition (working / not in use / repair) on the Assets screen")

    known_saps = {r[0].strip() for r in (await session.execute(
        select(inventory_t.c["SAP_Code"]))).all() if r[0]}
    known_saps |= extra_saps or set()   # dry-run: inventory was never written
    # Deliberately NOT filtered by site (alembic a3c17e9b25d4). Identity is
    # (SAP_Code, serial_no) globally, so a serial already registered at
    # ANOTHER site is an existing unit, not a new one — proposing it would
    # have the workbook conjure a second row for a hammer that already
    # exists elsewhere, and the insert would then hit the unique constraint
    # anyway. The site filter was correct only while site was part of the key.
    have = {(r[0], r[1]) for r in (await session.execute(
        select(asset_unit_t.c["SAP_Code"], asset_unit_t.c["serial_no"]))).all()}

    unknown_status = Counter()
    seen: set[tuple[str, str]] = set()
    for n, row in enumerate(rows, start=1):
        sap = _s(row[sap_i]) if sap_i is not None and sap_i < len(row) else None
        place = _s(row[loc_i]) if loc_i < len(row) else None
        if not sap:
            continue
        if not place:
            out["consumable_rows"] += 1     # THE GOLDEN RULE, negative half
            continue
        serial = _s(row[ser_i]) if ser_i < len(row) else None
        if not serial:
            out["no_serial"].append(
                {"row": n, "sap": sap, "location": place,
                 "date": _s(row[date_i]) if date_i is not None
                         and date_i < len(row) else None})
            continue
        if sap not in known_saps:
            out["rejects"].append({"row": n, "sap": sap,
                                   "reason": "SAP not in inventory master"})
            continue
        key = (sap, serial)
        if key in have:
            # APP WINS — no status, rack or fix is proposed. The Location TEXT
            # is offered separately, and `refresh_asset_location_notes` applies
            # it only while the app has never touched the unit.
            out["existing"] += 1
            out["seen_notes"].append({"SAP_Code": sap, "serial_no": serial,
                                      "location_note": place[:200]})
            continue
        if key in seen:
            out["duplicate_rows"] += 1
            continue
        seen.add(key)

        status = "in_stock"
        if st_i is not None and st_i < len(row):
            raw = _s(row[st_i])
            if raw:
                canon = _ASSET_STATUS_CANON.get(raw.strip().lower())
                if canon:
                    status = canon
                else:
                    unknown_status[raw] += 1
        out["inserts"].append({
            "Site_ID": site_id, "SAP_Code": sap, "serial_no": serial,
            "status": status,
            # Free text, not a rack id: the workbook says "At site", which is a
            # place a person recognises and not a shelf in `storage_locations`.
            "location_note": place[:200],
            "created_by": "excel-sync", "last_seen_by": "excel-sync",
        })

    if out["no_serial"]:
        sample = ", ".join(f"row {r['row']} (SAP {r['sap']}, {r['location']!r})"
                           for r in out["no_serial"][:5])
        out["warnings"].append(
            f"{len(out['no_serial'])} row(s) have a Location but no Serial No. "
            f"— cannot be keyed, so no asset was created: {sample}"
            + (" …" if len(out["no_serial"]) > 5 else ""))
    if unknown_status:
        out["warnings"].append(
            "unrecognised Status value(s), left as 'in_stock': "
            + ", ".join(f"{k!r} ×{v}" for k, v in unknown_status.items()))
    if out["existing"]:
        out["warnings"].append(
            f"{out['existing']} asset row(s) already exist — left untouched "
            f"(the app owns an asset's status, rack and GPS fix)")
    return out


async def apply_asset_units(session: AsyncSession, plan: dict,
                            username: str) -> dict:
    """Create the new units, each with its opening movement row.

    The registration IS the first movement — the same contract
    `assets.create_asset` keeps — so "where has this been" has no gap at the
    start of the history.

    THE APP WINS is enforced twice over: `plan_asset_units` never proposes a
    unit that already exists, and `DO NOTHING` here means that even a racing
    second run cannot overwrite a status or a GPS fix.
    """
    inserted = 0
    for unit in plan.get("inserts", []):
        new_id = (await session.execute(
            pg_insert(asset_unit_t).values(**unit, last_seen_at=func.now())
            # (SAP_Code, serial_no) — GLOBAL since alembic a3c17e9b25d4. The
            # constraint used to include Site_ID; naming the old one here is
            # what broke when the key was narrowed, because ON CONFLICT ON
            # CONSTRAINT fails hard on a name that no longer exists rather
            # than degrading. Worth knowing for the SEMANTICS too: a serial
            # already registered at ANOTHER site now conflicts and is skipped,
            # which is right — the workbook must not conjure a second row for
            # a hammer that already exists somewhere else.
            .on_conflict_do_nothing(constraint="uq_asset_units_sap_serial")
            .returning(asset_unit_t.c["id"]))).scalar()
        if new_id is None:
            continue                # already there — see the docstring
        await session.execute(insert(asset_move_t).values(
            asset_unit_id=new_id, moved_by=username,
            to_note=unit["location_note"], source="excel-sync",
            status=unit["status"], note="seeded from CNCEC_Inventory.xlsx"))
        inserted += 1
    if inserted:
        await write_audit(session, username, "BULK_IMPORT_ASSETS", "asset_units",
                          f"pg_excel_sync: +{inserted} asset unit(s) from the "
                          f"Consumption Log (Location ⇒ reusable asset)")
    return {"units": inserted}


async def refresh_asset_location_notes(session: AsyncSession, plan: dict,
                                       site_id: str) -> int:
    """Re-seed the Location TEXT on units the app has never touched.

    The narrow case the ruling leaves open: a unit this sync created on an
    earlier run, that nobody has since moved, whose spreadsheet cell has been
    corrected. Updating it is right — nothing in the app is being overwritten.

    The guard is `last_seen_by = 'excel-sync'` AND no coordinates: every app
    path (`create_asset`, `move_asset`) stamps `last_seen_by` with the real
    username, so this predicate is false the moment a human is involved. GPS is
    checked as well because a fix is the strongest possible statement that
    somebody stood next to this thing.
    """
    touched = 0
    for unit in plan.get("seen_notes", []):
        res = await session.execute(
            update(asset_unit_t)
            .where(asset_unit_t.c["Site_ID"] == site_id,
                   asset_unit_t.c["SAP_Code"] == unit["SAP_Code"],
                   asset_unit_t.c["serial_no"] == unit["serial_no"],
                   asset_unit_t.c["last_seen_by"] == "excel-sync",
                   asset_unit_t.c["current_lat"].is_(None),
                   asset_unit_t.c["current_lng"].is_(None),
                   asset_unit_t.c["location_note"].isnot(None),
                   asset_unit_t.c["location_note"] != unit["location_note"])
            .values(location_note=unit["location_note"]))
        touched += res.rowcount or 0
    return touched


# ─── THE APP WINS: operator SQM overrides survive every sync ──────────────────
#
# `sme_equipment.Surface_Area_SQM` drives DEMAND, so a UI correction that a
# workbook sync quietly reverted would show up a week later as a wrong buy list
# with nothing to point at. Ruling 2026-08-04: an override beats the workbook,
# and the divergence is REPORTED on every run rather than resolved in silence.
_EQ_KEY = ("Site_ID", "Equipment_Tag_No", "Lining_System_Code")


async def snapshot_sqm_overrides(session: AsyncSession, site_id: str) -> dict:
    """(site, tag, code) → the override row, captured BEFORE a `--sme-reseed`
    deletes it. `restore_sqm_overrides` puts it back afterwards."""
    e = equipment_t.c
    rows = (await session.execute(
        select(e["Site_ID"], e["Equipment_Tag_No"], e["Lining_System_Code"],
               e["SQM_Override"], e["SQM_Override_By"], e["SQM_Override_At"])
        .where(e["Site_ID"] == site_id)
        .where(e["SQM_Override"].isnot(None)))).mappings().all()
    return {(r["Site_ID"], r["Equipment_Tag_No"], r["Lining_System_Code"]): dict(r)
            for r in rows}


async def restore_sqm_overrides(session: AsyncSession, site_id: str,
                                snapshot: dict | None = None) -> list[dict]:
    """Re-apply overrides on top of whatever the workbook just wrote.

    Returns one entry per row where the workbook and the operator disagree, so
    the caller can print it. An override that MATCHES the workbook is silently
    satisfied — there is nothing to warn about.
    """
    e = equipment_t.c
    for key, snap in (snapshot or {}).items():
        await session.execute(
            update(equipment_t)
            .where(e["Site_ID"] == key[0])
            .where(e["Equipment_Tag_No"] == key[1])
            .where(e["Lining_System_Code"] == key[2])
            .values(SQM_Override=snap["SQM_Override"],
                    SQM_Override_By=snap["SQM_Override_By"],
                    SQM_Override_At=snap["SQM_Override_At"]))

    rows = (await session.execute(
        select(e["id"], e["Equipment_Tag_No"], e["Lining_System_Code"],
               e["Surface_Area_SQM"], e["SQM_Override"], e["SQM_Override_By"])
        .where(e["Site_ID"] == site_id)
        .where(e["SQM_Override"].isnot(None)))).mappings().all()
    diverged = []
    for r in rows:
        if float(r["Surface_Area_SQM"] or 0) == float(r["SQM_Override"]):
            continue
        diverged.append({"tag": r["Equipment_Tag_No"],
                         "code": r["Lining_System_Code"],
                         "workbook": float(r["Surface_Area_SQM"] or 0),
                         "override": float(r["SQM_Override"]),
                         "by": r["SQM_Override_By"]})
        await session.execute(update(equipment_t)
                              .where(e["id"] == r["id"])
                              .values(Surface_Area_SQM=r["SQM_Override"]))
    return diverged


async def apply_ledger(session: AsyncSession, plan: dict, username: str) -> None:
    for kind, spec in _LEDGER_SHEETS.items():
        section = plan["sections"].get(kind) or {}
        table = spec["table"]
        for row in section.get("inserts", []):
            await session.execute(insert(table).values(**row))
        for c in section.get("corrections", []):
            await session.execute(update(table).where(table.c["id"] == c["id"])
                                  .values(Quantity=c["qty_to"]))
        if section.get("inserts") or section.get("corrections"):
            await write_audit(session, username, "BULK_IMPORT_LEDGER",
                              table.name,
                              f"+{len(section['inserts'])} rows, "
                              f"{len(section['corrections'])} qty corrections")


# ─── SME masters ──────────────────────────────────────────────────────────────
async def plan_sme_equipment(session: AsyncSession, data: bytes, site_id: str) -> dict:
    headers, rows = _sheet_rows(data, "Data Input",
                                ("equipment_tag_no.", "lining_system_code"),
                                required=False)
    if not headers:
        headers, rows = _sheet_rows(data, None,
                                    ("equipment_tag_no.", "lining_system_code"))
    field_names = {  # workbook header → sme_equipment column
        "Sl. #": "Sl_No", "Project": "Project", "WBS #": "WBS_No",
        "IO#": "IO_No", "Sub_Location": "Sub_Location", "Location": "Location",
        "Type": "Type", "Substrate": "Substrate", "Name": "Name",
        "Drawing #": "Drawing_No", "Design": "Design", "Dia / L": "Dia_L",
        "Ht. /W": "Ht_W", "Equipment Total SQM": "Equipment_Total_SQM",
        "Remaraks": "Remaraks",
        "Lining_System_Short_Name": "Lining_System_Short_Name",
        "Lining_Type": "Lining_Type", "Lining_System": "Lining_System",
        "Material Spec.": "Material_Spec",
        "Lining_Area/location": "Lining_Area_Location",
    }
    ix = {col: _col(headers, hdr) for hdr, col in field_names.items()}
    tag_i = _col(headers, "Equipment_Tag_No.", "Equipment_Tag_No")
    code_i = _col(headers, "Lining_System_Code")
    sqm_i = _col(headers, "Surface_Area_SQM")
    if tag_i is None or code_i is None or sqm_i is None:
        raise HTTPException(422, "Equipment sheet needs Equipment_Tag_No., "
                                 "Lining_System_Code and Surface_Area_SQM")

    # short-name → code backfill map (recipes already in the DB)
    sn_map = {}
    for r in (await session.execute(
            select(recipe_t.c["Lining_System_Name"],
                   recipe_t.c["Lining_System_Code"]))).all():
        if r[0] and r[0].strip() and r[0].strip() not in sn_map:
            sn_map[r[0].strip()] = str(r[1]).strip()

    agg: dict[tuple[str, str], dict] = {}
    warnings, skipped_placeholder, backfilled_tags = [], 0, 0
    for row in rows:
        def cell(i):
            return row[i] if i is not None and i < len(row) else None
        tag = _s(cell(tag_i))
        name = _s(cell(ix["Name"]))
        if not tag and name:
            tag, backfilled_tags = name, backfilled_tags + 1  # Name IS the identity
        code = _s(cell(code_i))
        short = _s(cell(ix["Lining_System_Short_Name"]))
        if not code and short and short in sn_map:
            code = sn_map[short]
        sqm = _f(cell(sqm_i))
        if not tag or not code:
            continue
        if _is_placeholder_code(code):
            skipped_placeholder += 1
            continue
        if sqm is None or sqm <= 0:
            continue
        a = agg.setdefault((tag, code), {"Surface_Area_SQM": 0.0, "_areas": []})
        a["Surface_Area_SQM"] += sqm
        area = _s(cell(ix["Lining_Area_Location"]))
        if area and area not in a["_areas"]:
            a["_areas"].append(area)
        for col, i in ix.items():
            if col == "Lining_Area_Location":
                continue
            v = cell(i)
            if col == "Equipment_Total_SQM":
                v = _f(v)
            else:
                v = _s(v)
                if col == "Location" and v:
                    v = _LOCATION_CANON.get(v.lower(), v)
            if v is not None and col not in a:
                a[col] = v
    if skipped_placeholder:
        warnings.append(f"skipped {skipped_placeholder} row(s) whose "
                        f"Lining_System_Code is a placeholder (e.g. "
                        f"To_Be_Confirmed_LSC)")
    if backfilled_tags:
        warnings.append(f"backfilled Equipment_Tag_No from Name for "
                        f"{backfilled_tags} area row(s)")

    existing = {(r["Equipment_Tag_No"], r["Lining_System_Code"]): dict(r)
                for r in (await session.execute(
                    select(equipment_t).where(equipment_t.c["Site_ID"] == site_id)
                )).mappings().all()}
    inserts, updates, unchanged = [], [], 0
    for (tag, code), a in agg.items():
        areas = ", ".join(a.pop("_areas", [])) or None
        fields = {k: v for k, v in a.items() if v is not None}
        if areas:
            fields["Lining_Area_Location"] = areas
        fields["Surface_Area_SQM"] = round(float(fields["Surface_Area_SQM"]), 4)
        cur = existing.get((tag, code))
        if cur is None:
            inserts.append({"Site_ID": site_id, "Equipment_Tag_No": tag,
                            "Lining_System_Code": code, **fields})
        else:
            diff = {k: v for k, v in fields.items() if cur.get(k) != v}
            if diff:
                updates.append({"id": cur["id"], "tag": tag, "code": code,
                                "diff": diff,
                                "sqm": fields["Surface_Area_SQM"]})
            else:
                unchanged += 1
    not_in_file = [f"{t}/{c}" for (t, c) in existing if (t, c) not in agg]
    if not_in_file:
        warnings.append(f"{len(not_in_file)} DB equipment row(s) not in the file "
                        f"— left untouched (delete via Master Data if intended)")
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": [], "warnings": warnings, "site_id": site_id}


async def apply_sme_equipment(session: AsyncSession, plan: dict, username: str) -> None:
    site = plan["site_id"]
    for row in plan["inserts"]:
        await session.execute(insert(equipment_t).values(**row))
        await _upsert_progress(session, site, row["Equipment_Tag_No"],
                               row["Lining_System_Code"],
                               original_sqm=row["Surface_Area_SQM"], done_sqm=None)
    for u in plan["updates"]:
        await session.execute(update(equipment_t)
                              .where(equipment_t.c["id"] == u["id"])
                              .values(**u["diff"]))
        # re-seed the baseline, PRESERVING Done_SQM (legacy bootstrap contract)
        await _upsert_progress(session, site, u["tag"], u["code"],
                               original_sqm=u["sqm"], done_sqm=None)
    await write_audit(session, username, "BULK_IMPORT_SME_EQUIPMENT",
                      "sme_equipment",
                      f"{site}: +{len(plan['inserts'])} ~{len(plan['updates'])}")


async def plan_sme_recipes(session: AsyncSession, data: bytes) -> dict:
    headers, rows = _sheet_rows(data, None, ("lining_system_code", "material_code"))
    cols = {"Lining_System_Name": ("Lining_System_Short_Name",),
            "Substrate": ("Substrate",), "System_Keys": ("System Key's", "System_Keys"),
            "Lining_Thickness": ("Lining_Thicknes", "Lining_Thickness"),
            "Lining_System": ("Lining_System",), "Lining_Type": ("Lining_Type",),
            "Material_Description": ("Material_Description",),
            "Material_Name": ("Material_Name",), "UOM": ("UOM",),
            "Package_Size": ("PACKAGE SIZE", "Package_Size"),
            "Sl_No": ("Sl. #",)}
    ix = {field: _col(headers, *names) for field, names in cols.items()}
    code_i = _col(headers, "Lining_System_Code")
    mat_i = _col(headers, "Material_Code")
    sqm_i = _col(headers, "For_1_SQM")
    sap_i = _col(headers, "SAP_Code", "SAP CODE")
    esc_i = _col(headers, "Execution_Sub_Activity_Code", "Execution Sub Activity Code")
    sap_aware = sap_i is not None  # 2026-07-18 workbook layout
    esc_aware = esc_i is not None  # 2026-08 workbook layout
    if code_i is None or mat_i is None or sqm_i is None:
        raise HTTPException(422, "recipe sheet needs Lining_System_Code, "
                                 "Material_Code and For_1_SQM")
    existing = {(str(r["Lining_System_Code"]).strip(),
                 _s(r.get("Execution_Sub_Activity_Code")) or "",
                 r["Material_Code"], _s(r.get("SAP_Code")) or ""): dict(r)
                for r in (await session.execute(select(recipe_t))).mappings().all()}
    # Rows a sync has never classified: identity (code, '', material, SAP). The
    # ESC migration gave every pre-existing row '' because the codes live only
    # in the workbook. The FIRST sub-activity to claim one ADOPTS it (an update
    # that fills the ESC in place) rather than inserting a duplicate beside it;
    # its siblings insert normally. Without this, a system whose merged row was
    # split into a primer and a screed line would end up holding all three.
    unclassified = {k: v for k, v in existing.items() if k[1] == ""}
    adopted: set = set()
    rejects: list[dict] = []
    # Line identity is (code, SUB-ACTIVITY, material, SAP). PU systems carry
    # Comp-A/B/C/D lines that share a Material_Code and differ only by variant
    # SAP; a repeat of the SAME identity in a SAP-aware file is a deliberate
    # coat line and For_1_SQM sums. ⚠️ Before the ESC column existed that merge
    # also swallowed the primer/screed split — LSC2 Resin A is 0.2700 under
    # ESC21 and 1.4674 under ESC22, and the three-part key summed them to
    # 1.7374. With ESC in the key those are two lines, which is the point.
    # Legacy files (no SAP column) keep first-occurrence-wins dedupe.
    agg: dict[tuple, dict] = {}
    dup_skips, coat_merges = 0, 0
    for n, row in enumerate(rows, start=1):
        code, mat_cell = _s(row[code_i]), _s(row[mat_i])
        if not code or not mat_cell:
            continue
        if _is_placeholder_code(code):
            rejects.append({"row": n, "reason": f"placeholder code {code!r}"})
            continue
        sap = _s(row[sap_i]) if sap_aware and sap_i < len(row) else None
        esc = (_s(row[esc_i]) or "") if esc_aware and esc_i < len(row) else ""
        qty = _f(row[sqm_i]) or 0.0
        fields = {}
        for field, i in ix.items():
            v = _s(row[i]) if i is not None and i < len(row) else None
            if v is not None:
                fields[field] = v
        # a comma-separated Material_Code cell is one line per material
        for mat in (m.strip() for m in mat_cell.split(",")):
            if not mat:
                continue
            key = (code, esc, mat, sap or "")
            cur = agg.get(key)
            if cur is not None:
                if sap_aware:
                    cur["For_1_SQM"] += qty
                    coat_merges += 1
                else:
                    dup_skips += 1
                continue
            agg[key] = {"For_1_SQM": qty, **fields,
                        **({"SAP_Code": sap} if sap else {}),
                        "Execution_Sub_Activity_Code": esc}

    inserts, updates, unchanged, adoptions = [], [], 0, 0
    for (code, esc, mat, sap), fields in agg.items():
        cur = existing.get((code, esc, mat, sap))
        if cur is None and esc:
            # nothing at this identity — adopt this system/material/SAP's
            # still-unclassified row, if it has one and nobody took it yet.
            legacy_key = (code, "", mat, sap)
            legacy_row = unclassified.get(legacy_key)
            if legacy_row is not None and legacy_key not in adopted:
                cur = legacy_row
                adopted.add(legacy_key)
                adoptions += 1
        if cur is None:
            inserts.append({"Lining_System_Code": code, "Material_Code": mat,
                            **fields})
        else:
            diff = {k: v for k, v in fields.items() if cur.get(k) != v}
            if diff:
                updates.append({"id": cur["id"], "diff": diff})
            else:
                unchanged += 1
    warnings = []
    if dup_skips:
        warnings.append(f"{dup_skips} repeated (code, material) line(s) skipped "
                        f"— first occurrence wins (legacy bootstrap rule)")
    if coat_merges:
        warnings.append(f"{coat_merges} repeated (code, sub-activity, material, "
                        f"SAP) coat line(s) merged — For_1_SQM summed")
    if adoptions:
        warnings.append(f"{adoptions} previously unclassified line(s) adopted "
                        f"into a sub-activity (Execution_Sub_Activity_Code "
                        f"filled in place, not duplicated)")
    # Anything still '' after this sync is a line the workbook no longer
    # describes. Reported, never deleted: a recipe row is master data and the
    # sync is not the place to decide it is obsolete.
    if esc_aware:
        orphans = [k for k in unclassified if k not in adopted]
        if orphans:
            warnings.append(
                f"{len(orphans)} recipe line(s) remain unclassified — the "
                f"workbook names no sub-activity for them, e.g. "
                f"{', '.join('/'.join(x for x in o if x) for o in orphans[:3])}")
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": rejects, "warnings": warnings}


async def apply_sme_recipes(session: AsyncSession, plan: dict, username: str) -> None:
    for row in plan["inserts"]:
        await session.execute(insert(recipe_t).values(**row))
    for u in plan["updates"]:
        await session.execute(update(recipe_t).where(recipe_t.c["id"] == u["id"])
                              .values(**u["diff"]))
    await write_audit(session, username, "BULK_IMPORT_SME_RECIPES", "sme_recipe",
                      f"+{len(plan['inserts'])} ~{len(plan['updates'])}")


# ─── manpower norms (Manpower_Hour_Details.xlsx, Block A) ────────────────────
# The nine role COLUMNS of the workbook, mapped to mh_roles.Role_Code. Header
# spellings are matched case-insensitively and are exactly what the sheet ships
# (note the lower-case 'mortar mixer' and 'brick cutter' — the workbook is not
# consistent and normalising it here is cheaper than asking for a re-type).
_NORM_ROLE_COLUMNS = [
    ("BLASTER", "Blaster"), ("POTMAN", "Potman"),
    ("RUBBER_LINER", "Rubber Liner"), ("COATING_APPLICATOR", "Coating applicator"),
    ("SHEET_PREPARATOR", "Sheet Preparator"), ("MASON", "Mason"),
    ("MORTAR_MIXER", "mortar mixer"), ("BRICK_CUTTER", "brick cutter"),
    ("HELPER", "Helper"),
]

# The measured fields. Two rows that agree on the identity AND on every one of
# these are the same benchmark written twice; two that disagree on any of them
# are a real collision the operator has to resolve.
_NORM_VALUE_FIELDS = ("Crew_Size", "Hours_Per_Shift", "Manhours_Per_Shift",
                      "Standard_Productivity_Per_Shift", "SQM_Per_Hour_Per_Person")


async def plan_sme_manpower_norms(session: AsyncSession, data: bytes) -> dict:
    """Plan Block A of Manpower_Hour_Details.xlsx into sme_manpower_norm.

    ⚠️ BLOCK B IS NOT READ. Rows 41-49 are a worked day/night example, not
    master data (operator ruling, 2026-08-18). They are excluded structurally
    rather than by row number: every Block A row names a `Type` and a
    `Lining_System_Code` and no Block B row does, so a sheet that grows or
    shifts does not silently start importing the example.

    ⚠️ A COLLISION IS REJECTED, NOT MERGED. Where two rows share the identity
    but disagree on the numbers, keeping either one plans a crew against a
    benchmark that can be 7.5x wrong. The reject names the exact `Variant_Key`
    to type, because "duplicate row" on its own is not actionable.
    """
    headers, rows = _sheet_rows(data, "Productivity Estimation",
                                ("lining_system_code", "activity"),
                                required=False)
    if not headers:
        headers, rows = _sheet_rows(data, None,
                                    ("lining_system_code", "activity"))
    ix = {
        "Activity_Code": _col(headers, "Activity Code#", "Activity_Code"),
        "Type": _col(headers, "Type"),
        "System": _col(headers, "System"),
        "Lining_System_Code": _col(headers, "Lining_System_Code"),
        "Activity": _col(headers, "Activity"),
        "Execution_Sub_Activity_Code": _col(headers, "Execution_Sub_Activity_Code"),
        "Sub_Activity": _col(headers, "Sub-Activity", "Sub_Activity"),
        "Variant_Key": _col(headers, "Variant_Key", "Variant Key"),
        "Crew_Size": _col(headers, " Person/Crew", "Person/Crew", "Person / Crew"),
        "Hours_Per_Shift": _col(headers, "Hrs./shift", "Hrs/shift"),
        "Manhours_Per_Shift": _col(headers, "Manhr. / Shift", "Manhr./Shift"),
        "Standard_Productivity_Per_Shift": _col(
            headers, "Standard Productivity /Shift", "Standard Productivity/Shift"),
        "SQM_Per_Hour_Per_Person": _col(headers, "SQ. Mtr/Hr./Person",
                                        "SQ.Mtr/Hr./Person"),
        "Remarks": _col(headers, "Remarks"),
    }
    for need in ("Type", "Lining_System_Code", "Execution_Sub_Activity_Code",
                 "Activity"):
        if ix[need] is None:
            raise HTTPException(422, f"manpower sheet needs a {need} column")
    role_ix = {code: _col(headers, name) for code, name in _NORM_ROLE_COLUMNS}

    def cell(row, i):
        return row[i] if i is not None and i < len(row) else None

    existing = {}
    for r in (await session.execute(select(norm_t))).mappings().all():
        existing[(_s(r["Type"]) or "", _s(r["Lining_System_Code"]) or "",
                  _s(r["Execution_Sub_Activity_Code"]) or "",
                  _s(r["Activity"]) or "", _s(r.get("Variant_Key")) or "")] = dict(r)

    agg: dict[tuple, dict] = {}
    crews: dict[tuple, dict] = {}
    rejects: list[dict] = []
    dup_skips = 0
    skipped_block_b = 0

    for n, row in enumerate(rows, start=1):
        typ = _s(cell(row, ix["Type"]))
        code = _s(cell(row, ix["Lining_System_Code"]))
        if not typ or not code:
            # Count only rows that HOLD something. openpyxl hands back the
            # sheet's trailing blank rows too, and folding those into the
            # figure turns a precise "19 example rows ignored" into an
            # alarming 62 that invites someone to go looking for data loss.
            if any(_s(v) for v in row):
                skipped_block_b += 1    # Block B, or a spacer row
            continue
        if _is_placeholder_code(code):
            rejects.append({"row": n, "reason": f"placeholder code {code!r}"})
            continue
        esc = _s(cell(row, ix["Execution_Sub_Activity_Code"]))
        act = _s(cell(row, ix["Activity"]))
        if not esc or not act:
            rejects.append({"row": n, "reason": "row names no sub-activity code "
                                                "or activity"})
            continue
        variant = _s(cell(row, ix["Variant_Key"])) or ""
        fields = {
            "Activity_Code": _s(cell(row, ix["Activity_Code"])),
            "Type": typ, "System": _s(cell(row, ix["System"])),
            "Lining_System_Code": code, "Execution_Sub_Activity_Code": esc,
            "Activity": act, "Sub_Activity": _s(cell(row, ix["Sub_Activity"])),
            "Variant_Key": variant,
            "Remarks": _s(cell(row, ix["Remarks"])),
        }
        for f in _NORM_VALUE_FIELDS:
            fields[f] = _f(cell(row, ix[f])) or 0.0
        crew = {rc: (_f(cell(row, i)) or 0.0)
                for rc, i in role_ix.items()
                if i is not None and (_f(cell(row, i)) or 0.0) > 0}

        key = (typ, code, esc, act, variant)
        prev = agg.get(key)
        if prev is not None:
            same = all(abs((prev.get(f) or 0.0) - (fields.get(f) or 0.0)) < 1e-9
                       for f in _NORM_VALUE_FIELDS) and crews.get(key) == crew
            if same:
                dup_skips += 1          # the identical repeat — benign
                continue
            rejects.append({
                "row": n,
                "reason": (
                    f"collision: {typ}/{code}/{esc}/{act!r} is already defined "
                    f"with different numbers (crew {prev.get('Crew_Size')} @ "
                    f"{prev.get('Standard_Productivity_Per_Shift')} /shift vs "
                    f"crew {fields.get('Crew_Size')} @ "
                    f"{fields.get('Standard_Productivity_Per_Shift')} /shift). "
                    f"Give the two rows different Activity text, or add a "
                    f"Variant_Key column and a distinct value in each"),
            })
            continue
        agg[key] = fields
        crews[key] = crew

    inserts, updates, unchanged = [], [], 0
    for key, fields in agg.items():
        cur = existing.get(key)
        if cur is None:
            inserts.append({**fields, "_crew": crews[key]})
        else:
            diff = {k: v for k, v in fields.items() if cur.get(k) != v}
            if diff:
                updates.append({"id": cur["id"], "diff": diff,
                                "_crew": crews[key]})
            else:
                unchanged += 1

    warnings = []
    if dup_skips:
        warnings.append(f"{dup_skips} identical repeat(s) of an existing "
                        f"benchmark collapsed — same identity, same numbers")
    if skipped_block_b:
        warnings.append(f"{skipped_block_b} non-blank row(s) skipped as "
                        f"not-a-benchmark (no Type / Lining_System_Code) — "
                        f"Block B's day/night worked example lives there and "
                        f"is deliberately not imported")
    unknown_roles = [c for c, i in role_ix.items() if i is None]
    if unknown_roles:
        warnings.append(f"no column found for role(s) {unknown_roles} — their "
                        f"headcounts will be absent from every crew")
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": rejects, "warnings": warnings}


async def apply_sme_manpower_norms(session: AsyncSession, plan: dict,
                                   username: str) -> None:
    for row in plan["inserts"]:
        crew = row.pop("_crew", {})
        new_id = (await session.execute(
            insert(norm_t).values(**row).returning(norm_t.c["id"]))).scalar_one()
        await _write_norm_crew(session, new_id, crew)
    for u in plan["updates"]:
        crew = u.pop("_crew", {})
        await session.execute(update(norm_t).where(norm_t.c["id"] == u["id"])
                              .values(**u["diff"]))
        await _write_norm_crew(session, u["id"], crew)
    await write_audit(session, username, "BULK_IMPORT_SME_MANPOWER",
                      "sme_manpower_norm",
                      f"+{len(plan['inserts'])} ~{len(plan['updates'])}")


async def _write_norm_crew(session: AsyncSession, norm_id: int,
                           crew: dict) -> None:
    """Replace a norm's crew wholesale.

    Deleted-then-written rather than upserted: a role REMOVED from the workbook
    has to disappear, and an upsert cannot express an absence.
    """
    await session.execute(delete(norm_role_t)
                          .where(norm_role_t.c["Norm_ID"] == norm_id))
    for role_code, headcount in crew.items():
        await session.execute(insert(norm_role_t).values(
            Norm_ID=norm_id, Role_Code=role_code, Headcount=headcount))


async def plan_sme_materials(session: AsyncSession, data: bytes) -> dict:
    headers, rows = _sheet_rows(data, None, ("material_code", "material_name"))
    ix = {"Item": _col(headers, "Item"),
          "Vendor": _col(headers, "Vendor/supplying plant", "Vendor"),
          "Purchasing_Document": _col(headers, "Purchasing Document"),
          "Document_Date": _col(headers, "Document Date"),
          "Material_Name": _col(headers, "Material_Name"),
          "Nature": _col(headers, "Nature"), "UOM": _col(headers, "UOM"),
          "sap": _col(headers, "SAP_Code", "SAP CODE"),
          # `_col` matches header names EXACTLY (case-insensitively), so every
          # spelling the workbook actually ships has to be listed. The live
          # Materials_DetailsAvailable_Qty.xlsx writes "Available Qty" with a
          # SPACE while the ordered column keeps its underscore; with only the
          # underscored alias here the availability column resolved to None and
          # every material silently re-baselined to 0 — 29 of them, on real
          # data, caught by a dry-run. Missing columns are now handled below
          # rather than defaulting to zero, but the aliases come first.
          "avail": _col(headers, "Available_Qty", "Available Qty",
                        "Available Quantity", "Available"),
          "ordered": _col(headers, "Ordered_Qty", "Ordered Qty",
                          "Ordered Quantity", "Ordered")}
    mat_i = _col(headers, "Material_Code")
    if mat_i is None:
        raise HTTPException(422, "Material_Code column missing")
    # A quantity column the workbook does not carry must LEAVE THE STORED VALUE
    # ALONE, never overwrite it with the 0.0 that summing no cells produces.
    # Omitting the field entirely is what makes that true on both write paths:
    # the upsert's COALESCE(excluded, table) keeps the stored value, and the
    # diff below can never propose a change for a field that is not there.
    qty_fields = [(k, f"Initial_{'Available' if k == 'avail' else 'Ordered'}_Qty")
                  for k in ("avail", "ordered") if ix[k] is not None]
    missing_qty = [k for k in ("avail", "ordered") if ix[k] is None]
    # 2026-07-30 COMPONENT IDENTITY: aggregate per (Material_Code, SAP_Code).
    # This used to key on Material_Code alone, which SUMMED the four Comp-A/B/C/D
    # rows of a PU system into one stock figure and joined their SAPs into a
    # comma list — four distinct drums recorded as one bucket. Repeat rows for
    # the SAME component (several purchase documents for one SAP) still sum,
    # which is the aggregation that was always intended.
    agg: dict[tuple[str, str], dict] = {}
    for row in rows:
        mat = _s(row[mat_i]) if mat_i < len(row) else None
        if not mat:
            continue

        def cell(key):
            i = ix[key]
            return row[i] if i is not None and i < len(row) else None
        # The ERP writes the same variant as "1043-2" and "1043 - 2".
        sap = "".join((_s(cell("sap")) or "").split())
        a = agg.setdefault((mat, sap), {col: 0.0 for _, col in qty_fields})
        for key, col in qty_fields:
            a[col] += _f(cell(key)) or 0.0
        dd = cell("Document_Date")
        dd = (_iso(dd) or "")[:10] or None
        if dd and dd > (a.get("Document_Date") or ""):
            a["Document_Date"] = dd  # most recent PO date wins
        for field in ("Item", "Vendor", "Purchasing_Document",
                      "Material_Name", "Nature", "UOM"):
            v = _s(cell(field))
            if v is not None and field not in a:
                a[field] = v
    existing = {(r["Material_Code"], _s(r.get("SAP_Code")) or ""): dict(r)
                for r in (await session.execute(select(seed_t))).mappings().all()}
    inserts, updates, unchanged = [], [], 0
    for (mat, sap), a in agg.items():
        for _, col in qty_fields:
            a[col] = round(a[col], 4)
        cur = existing.get((mat, sap))
        if cur is None:
            inserts.append({"Material_Code": mat, "SAP_Code": sap, **a})
        else:
            diff = {k: v for k, v in a.items() if cur.get(k) != v}
            if diff:
                updates.append({"Material_Code": mat, "SAP_Code": sap, "diff": diff})
            else:
                unchanged += 1
    # ── retire SAP-less placeholder rows the workbook has superseded ─────────
    # The FROZEN legacy SQLite seed has no SAP_Code column at all, so a cutover
    # lands every material as one row with SAP_Code = ''. The first workbook sync
    # then inserts the real per-component rows beside it, and without this the
    # placeholder lingers forever — a phantom material carrying the whole
    # pre-split quantity, double-counting the stock it was replaced by.
    #
    # Scoped tightly, on TWO conditions:
    #   1. the workbook supplied at least one real SAP for that Material_Code
    #      (a material the workbook genuinely lists without one keeps its row);
    #   2. no SAP-LESS RECIPE line still references that material. The frozen
    #      legacy DB carries 86 pre-workbook recipe rows with no SAP at all, and
    #      those lines can only draw on a blank-SAP seed row — retiring it would
    #      leave them with a zero pool and read as a total shortage. When this
    #      guard holds a row back, the fix is the documented SME reseed
    #      (`pg_excel_sync --sme-reseed`), which replaces both sides from the
    #      workbook, NOT deleting live stock rows out from under a live recipe.
    coded = {mat for (mat, sap) in agg if sap}
    sapless_recipe_mats = {
        r[0] for r in (await session.execute(
            select(recipe_t.c["Material_Code"])
            .where(func.coalesce(func.trim(recipe_t.c["SAP_Code"]), "") == ""))).all()}
    stale = [{"Material_Code": mat, "SAP_Code": sap}
             for (mat, sap) in existing
             if mat in coded and not sap and (mat, sap) not in agg
             and mat not in sapless_recipe_mats]
    held = sorted({mat for (mat, sap) in existing
                   if mat in coded and not sap and (mat, sap) not in agg
                   and mat in sapless_recipe_mats})
    warnings = []
    if missing_qty:
        warnings.append(
            "no "
            + " or ".join("Available_Qty" if k == "avail" else "Ordered_Qty"
                          for k in missing_qty)
            + f" column in this workbook (headers: {', '.join(headers)}) — "
            f"those quantities are LEFT AS THEY ARE in the database rather "
            f"than re-baselined to 0. Rename the column if that is not what "
            f"you meant.")
    if stale:
        warnings.append(
            f"{len(stale)} SAP-less placeholder row(s) superseded by the "
            f"workbook's per-component rows will be removed "
            f"({', '.join(sorted(s['Material_Code'] for s in stale)[:5])}"
            f"{'…' if len(stale) > 5 else ''})")
    if held:
        warnings.append(
            f"{len(held)} pooled seed row(s) KEPT because SAP-less recipe lines "
            f"still reference them — the stock would otherwise vanish from under "
            f"a live recipe. Run an SME reseed (--sme-reseed) to replace both "
            f"sides from the workbook: "
            f"{', '.join(held[:5])}{'…' if len(held) > 5 else ''}")
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": [], "warnings": warnings, "stale": stale}


async def apply_sme_materials(session: AsyncSession, plan: dict, username: str) -> None:
    keys = ("Material_Code", "SAP_Code")
    for row in plan["inserts"]:
        stmt = pg_insert(seed_t).values(**row, updated_at=func.now())
        stmt = stmt.on_conflict_do_update(
            index_elements=list(keys),
            set_={**{k: stmt.excluded[k] for k in row if k not in keys},
                  "updated_at": func.now()})
        await session.execute(stmt)
    for u in plan["updates"]:
        await session.execute(update(seed_t)
                              .where(seed_t.c["Material_Code"] == u["Material_Code"],
                                     seed_t.c["SAP_Code"] == u["SAP_Code"])
                              .values(**u["diff"], updated_at=func.now()))
    stale = plan.get("stale") or []
    for s in stale:
        await session.execute(delete(seed_t).where(
            seed_t.c["Material_Code"] == s["Material_Code"],
            seed_t.c["SAP_Code"] == s["SAP_Code"]))
    await write_audit(session, username, "BULK_IMPORT_SME_MATERIALS",
                      "sme_inventory_seed",
                      f"+{len(plan['inserts'])} ~{len(plan['updates'])} "
                      f"-{len(stale)}")


# ─── endpoints ────────────────────────────────────────────────────────────────
def _summary(plan: dict) -> dict:
    if "sections" in plan:  # ledger
        return {k: {"inserts": len(s["inserts"]),
                    "corrections": len(s["corrections"]),
                    "matched": s["matched"], "zero_skipped": s["zero_skipped"],
                    "db_only": s["db_only"]}
                for k, s in plan["sections"].items()}
    out = {"inserts": len(plan["inserts"]), "updates": len(plan["updates"]),
           "unchanged": plan["unchanged"], "rejects": len(plan["rejects"])}
    if plan.get("stale"):
        out["stale_removed"] = len(plan["stale"])
    return out


async def _read_upload(file: UploadFile) -> bytes:
    data = await file.read()
    if len(data) > MAX_XLSX_BYTES:
        raise HTTPException(413, "workbook exceeds the 8 MB import cap")
    if not data[:4] == b"PK\x03\x04":
        raise HTTPException(422, "expected an .xlsx upload")
    return data


_SME_KINDS = {
    "sme-equipment": (plan_sme_equipment, apply_sme_equipment, True),
    "sme-recipes": (plan_sme_recipes, apply_sme_recipes, False),
    "sme-materials": (plan_sme_materials, apply_sme_materials, False),
}


@router.post("/{kind}", summary="Dry-run (default) or commit a bulk Excel import")
async def bulk_import(kind: str,
                      file: UploadFile = File(...),
                      commit: bool = Query(False),
                      site_id: Optional[str] = Query(None),
                      user: dict = Depends(require_roles("hod")),
                      session: AsyncSession = Depends(get_session)):
    data = await _read_upload(file)
    if kind in ("inventory", "ledger"):
        if user["level"] < 4:
            raise HTTPException(403, "inventory/ledger import is admin-only")
        site = (site_id or "CNCEC").strip()
        plan = await (plan_inventory(session, data, site) if kind == "inventory"
                      else plan_ledger(session, data, site))
        if commit:
            await (apply_inventory if kind == "inventory"
                   else apply_ledger)(session, plan, user["username"])
            await session.commit()
    elif kind in _SME_KINDS:
        plan_fn, apply_fn, scoped = _SME_KINDS[kind]
        if scoped:
            site = _write_site(user, site_id)
            plan = await plan_fn(session, data, site)
        else:
            plan = await plan_fn(session, data)
        if commit:
            await apply_fn(session, plan, user["username"])
            await session.commit()
    else:
        raise HTTPException(404, f"unknown import kind {kind!r} (use one of "
                                 f"inventory, ledger, sme-equipment, "
                                 f"sme-recipes, sme-materials)")
    resp = {"kind": kind, "committed": bool(commit), "summary": _summary(plan),
            "warnings": plan.get("warnings", []),
            "rejects": plan.get("rejects", [])[:200]}
    if not commit:  # preview payload for the UI (trimmed)
        if "sections" in plan:
            resp["preview"] = {k: {"inserts": s["inserts"][:20],
                                   "corrections": s["corrections"][:20]}
                               for k, s in plan["sections"].items()}
        else:
            resp["preview"] = {"inserts": plan["inserts"][:20],
                               "updates": plan["updates"][:20]}
    return resp
