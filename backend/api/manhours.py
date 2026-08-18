"""
backend/api/manhours.py — Man-Hour & Labor Tracking portal (Phase-10 parity).

Async PG port of the legacy manhour portal (pages_internal/manhour_portal.py +
the mh_* helpers in root database.py). Exact-locked to {hod, admin} via
require_roles("hod") — the same lock as the legacy page and the SME estimator.
HOD accounts (level 2 < SITE_SCOPE_MIN_LEVEL) are pinned to their own Site_ID;
admins pass ?site_id= (required on writes, optional on reads).

Isolation contract (unchanged from legacy): WRITES only mh_* tables; READS
sme_equipment / sme_recipe read-only for the Tag/Location/System dropdowns.
Never touches the material ledger or any sme_* write. ZERO new tables.

Hour math (ported verbatim): Total = (Out − In) − break, overnight wraps +24h;
Normal = min(Total, 8); OT = remainder. The attendance workbook's own dirty
hour columns are ignored — hours are always recomputed from In/Out.

The Estimate-vs-Actual endpoint inlines the legacy v_mh_estimate_vs_actual
view as plain SQL (the PG schema has no view — keeping it a query means no
migration). Exports reuse the shared /reports renderers (DRY).
"""
from __future__ import annotations

import datetime as _dt
import io
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, select, text, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from .auth import require_roles, resolve_site_param, site_scope
from .db import get_session
from .services.ledger import _MD, write_audit

employees_t = _MD.tables["mh_employees"]
# QSEP — the PERSON registry, mirrored into by the attendance import so the
# roster and the SMR worker list finally agree. `employees_t` above is the
# per-site EMPLOYMENT record; these are different tables and the names are
# spelled out here because confusing them is the bug this fixed.
employees_t_master = _MD.tables["employees"]
timesheets_t = _MD.tables["mh_timesheets"]
estimates_t = _MD.tables["mh_manhour_estimates"]
production_t = _MD.tables["mh_production"]
variance_t = _MD.tables["mh_variance_notes"]
sme_equipment_t = _MD.tables["sme_equipment"]
sme_recipe_t = _MD.tables["sme_recipe"]

router = APIRouter(prefix="/mh", tags=["man-hours"],
                   dependencies=[Depends(require_roles("hod"))])

# Overtime begins after this many NET worked hours. Two thresholds, because a
# GI employee and a supplied worker are on different contracts — 11 hours worked
# is 8 normal + 3 OT for GI, and 10 normal + 1 OT for a non-GI worker.
#
# These are DEFAULTS, not the rule: an HOD sets the live values through
# /mh/settings (app_settings keys below). Hard-coding them would make a
# contract change a code change.
MH_OT_THRESHOLD_DEFAULTS = {"GI": 8.0, "NON_GI": 10.0}
MH_OT_SETTING_KEYS = {"GI": "mh_ot_threshold_gi",
                      "NON_GI": "mh_ot_threshold_non_gi"}
# Kept for the callers that have no worker in hand (an estimate, an import
# with no roster row). It is the stricter of the two on purpose: crediting OT
# that was not earned is the error that costs money and nobody reports.
MH_NORMAL_THRESHOLD_HOURS = MH_OT_THRESHOLD_DEFAULTS["GI"]
MH_DEFAULT_BREAK_MINS = 60

# The physical shift is 12 hours either way — 11 worked plus 1 hour of lunch.
# `Shift` records WHICH one a worker is on, never how long it is.
MH_SHIFTS = ("Day", "Night")
MH_WORKER_TYPES = ("GI", "NON_GI")

# Blank-ish markers normalized to NULL on every write path. 'nan' is the
# legacy pandas str(NaN) artifact that polluted the bootstrap import (fixed by
# a one-time UPDATE in both DBs on 2026-07-05); this guard keeps it out for good.
_BLANKISH = {"", "nan", "none", "null"}

# Legend from the attendance workbook's ADD EMPLOYEE sheet. The KEYS were
# renamed OWN→GI and Supply→NON_GI on 2026-08-18 (alembic a7e2c9d41b83); the
# VALUES are company names and are unchanged.
_COMPANY_DEFAULTS = {"GI": "GI", "NON_GI": "DMC"}

# What the attendance workbook writes in its `type` column, mapped onto the
# stored vocabulary. Kept because the workbook still ships the old words — the
# rename was ours, not the operator's.
_WORKER_TYPE_ALIASES = {"own": "GI", "gi": "GI",
                        "supply": "NON_GI", "non_gi": "NON_GI",
                        "non-gi": "NON_GI", "nongi": "NON_GI"}


def normalize_worker_type(v) -> Optional[str]:
    """Workbook/legacy spelling → stored vocabulary, or None if unrecognised.

    None is returned rather than a default: silently classifying a worker
    decides their overtime threshold by accident.
    """
    s = str(v or "").strip().lower().replace(" ", "_")
    return _WORKER_TYPE_ALIASES.get(s)


async def ot_thresholds(session: AsyncSession) -> dict:
    """{worker_type: net hours before OT}, HOD-configured or defaulted."""
    out = dict(MH_OT_THRESHOLD_DEFAULTS)
    rows = (await session.execute(text(
        "SELECT key, value FROM app_settings WHERE key = ANY(CAST(:k AS text[]))"),
        {"k": list(MH_OT_SETTING_KEYS.values())})).all()
    by_key = {str(k): v for k, v in rows}
    for wt, key in MH_OT_SETTING_KEYS.items():
        raw = by_key.get(key)
        try:
            if raw is not None and str(raw).strip() != "":
                out[wt] = float(raw)
        except (TypeError, ValueError):
            pass        # a corrupt setting falls back to the default, loudly
    return out


def _clean(v) -> Optional[str]:
    """Trimmed string, or None when empty/blank-ish ('nan', 'none', 'null')."""
    s = str(v or "").strip()
    return None if s.lower() in _BLANKISH else s


def _rows(res):
    return [dict(m) for m in res.mappings().all()]


def _write_site(user: dict, site_id: Optional[str]) -> str:
    """Site for a WRITE: scoped users (hod) are pinned to their own site and
    must have one; admins must say which site they are editing."""
    scope = site_scope(user)
    if scope is not None:
        if site_id is not None and site_id.strip() and site_id.strip() != scope:
            raise HTTPException(403, "you may only edit data for your own site")
        if not scope:
            raise HTTPException(403, "your account has no Site_ID bound")
        return scope
    sid = (site_id or "").strip()
    if not sid:
        raise HTTPException(422, "site_id is required")
    return sid


# --- hour computation (verbatim port of database.compute_mh_hours) -------------
def _time_to_minutes(value) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, _dt.time):
        return value.hour * 60 + value.minute
    parts = str(value).strip().split(":")
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        return h * 60 + m
    except (ValueError, IndexError):
        return None


async def thresholds_for_codes(session: AsyncSession, sid: str,
                               codes) -> dict:
    """{employee_code: OT threshold} for a whole batch — ONE query.

    Resolved per batch rather than per row: both timesheet writers are loops,
    and a per-row lookup turns a 300-line attendance import into 300 extra
    round-trips. An employee with no roster row is absent from the map and the
    caller falls back to the stricter GI threshold.
    """
    codes = [c for c in {str(c or "").strip() for c in codes} if c]
    thresholds = await ot_thresholds(session)
    if not codes:
        return {}
    rows = (await session.execute(
        select(employees_t.c["Employee_Code"], employees_t.c["Worker_Type"])
        .where(employees_t.c["Site_ID"] == sid,
               employees_t.c["Employee_Code"].in_(codes)))).all()
    return {str(c): thresholds.get(str(wt or ""), thresholds["GI"])
            for c, wt in rows}


def compute_mh_hours(in_time, out_time,
                     break_mins: int = MH_DEFAULT_BREAK_MINS,
                     threshold: float = MH_NORMAL_THRESHOLD_HOURS,
                     ) -> tuple[float, float, float]:
    """(total, normal, overtime) net hours.

    `threshold` is where overtime starts and depends on the WORKER, not the
    shift: both types work the same 12-hour shift (11 net), and it splits
    8 + 3 for GI and 10 + 1 for non-GI. Callers that know the worker pass
    their threshold; those that do not get the stricter GI default.
    """
    im, om = _time_to_minutes(in_time), _time_to_minutes(out_time)
    if im is None or om is None:
        return 0.0, 0.0, 0.0
    gross = om - im
    if gross < 0:
        gross += 24 * 60  # overnight shift guard
    net = max(0.0, (gross - int(break_mins or 0)) / 60.0)
    total = round(net, 2)
    thr = float(threshold if threshold is not None else MH_NORMAL_THRESHOLD_HOURS)
    normal = round(min(total, thr), 2)
    ot = round(max(0.0, total - thr), 2)
    return total, normal, ot


# --- dropdown metadata (READ-ONLY over the frozen sme_* tables) -----------------
@router.get("/meta", summary="Dropdowns: equipment tags (+locations) and system codes")
async def meta(site_id: Optional[str] = None,
               user: dict = Depends(require_roles("hod")),
               session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    e = sme_equipment_t
    stmt = select(e.c["Equipment_Tag_No"], e.c["Location"]).where(
        e.c["Equipment_Tag_No"].is_not(None))
    if sid is not None:
        stmt = stmt.where(e.c["Site_ID"] == sid)
    tag_loc: dict[str, str] = {}
    for tag, loc in (await session.execute(stmt.order_by(e.c["Equipment_Tag_No"]))).all():
        tag_loc.setdefault(str(tag), str(loc or ""))
    codes = (await session.execute(
        select(sme_recipe_t.c["Lining_System_Code"]).distinct()
        .where(sme_recipe_t.c["Lining_System_Code"].is_not(None))
        .order_by(sme_recipe_t.c["Lining_System_Code"]))).scalars().all()
    return {"equipment_tags": sorted(tag_loc), "tag_locations": tag_loc,
            "system_codes": [str(c) for c in codes]}


# --- Overtime settings (HOD-configurable, app-level) --------------------------
class OtSettingsIn(BaseModel):
    gi: Optional[float] = Field(default=None, ge=0, le=24)
    non_gi: Optional[float] = Field(default=None, ge=0, le=24)


@router.get("/settings", summary="Overtime thresholds (net hours before OT)")
async def get_ot_settings(user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    live = await ot_thresholds(session)
    return {"thresholds": live, "defaults": dict(MH_OT_THRESHOLD_DEFAULTS),
            "shift_hours_note": "The physical shift is 12 hours (11 worked + "
                                "1 hour lunch) for both Day and Night. These "
                                "thresholds split those worked hours into "
                                "normal and overtime.",
            "worker_types": list(MH_WORKER_TYPES)}


@router.put("/settings", summary="Set an overtime threshold")
async def put_ot_settings(body: OtSettingsIn = Body(...),
                          user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    """HOD-level, not admin-level, and deliberately so.

    The thresholds are a contract term the HOD owns; routing them through
    /admin/settings would mean the person accountable for the labour figures
    cannot change them without raising a ticket.

    ⚠️ Changing a threshold does NOT rewrite timesheets already posted. Hours
    are split at write time, and silently re-splitting history would move
    overtime somebody has already been paid for.
    """
    changes = {k: v for k, v in body.model_dump().items() if v is not None}
    if not changes:
        raise HTTPException(422, "nothing to set")
    written = {}
    for field, wt in (("gi", "GI"), ("non_gi", "NON_GI")):
        if field not in changes:
            continue
        key, val = MH_OT_SETTING_KEYS[wt], str(float(changes[field]))
        res = await session.execute(text(
            "UPDATE app_settings SET value = :v WHERE key = :k"),
            {"k": key, "v": val})
        if res.rowcount == 0:
            await session.execute(text(
                "INSERT INTO app_settings (key, value) VALUES (:k, :v) "
                "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value"),
                {"k": key, "v": val})
        written[wt] = float(val)
    await write_audit(session, user["username"], "MH_OT_THRESHOLD_UPDATE",
                      "app_settings",
                      " ".join(f"{k}={v}" for k, v in sorted(written.items())))
    await session.commit()
    return {"updated": True, "thresholds": await ot_thresholds(session)}


# --- Employees (labor roster — logically separate from the system users table) --
class EmployeeIn(BaseModel):
    employee_code: str
    name: str
    designation: Optional[str] = ""
    worker_type: str = "GI"   # GI | NON_GI (the workbook's OWN | Supply)
    shift: str = "Day"        # Day | Night
    company: Optional[str] = ""
    site_id: Optional[str] = None


@router.get("/employees", summary="Labor roster")
async def list_employees(site_id: Optional[str] = None, status: Optional[str] = None,
                         user: dict = Depends(require_roles("hod")),
                         session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    t = employees_t
    stmt = select(t.c["id"], t.c["Site_ID"], t.c["Employee_Code"], t.c["Name"],
                  t.c["Designation"], t.c["Worker_Type"], t.c["Shift"],
                  t.c["Company"], t.c["status"], t.c["created_at"])
    if sid is not None:
        stmt = stmt.where(t.c["Site_ID"] == sid)
    if status:
        stmt = stmt.where(t.c["status"] == status)
    rows = _rows(await session.execute(stmt.order_by(t.c["Employee_Code"])))
    # The roster is where somebody checks WHY a worker's overtime starts where
    # it does, so ship the threshold beside the type rather than making the
    # reader hold the settings page in their head.
    thresholds = await ot_thresholds(session)
    for r in rows:
        r["OT_After_Hours"] = thresholds.get(str(r.get("Worker_Type") or ""),
                                             thresholds["GI"])
    return {"items": rows, "ot_thresholds": thresholds,
            "worker_types": list(MH_WORKER_TYPES), "shifts": list(MH_SHIFTS)}


@router.post("/employees", summary="Add or update a roster row (upsert on Site+Code)")
async def upsert_employee(body: EmployeeIn = Body(...),
                          user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    sid = _write_site(user, body.site_id)
    code, name = body.employee_code.strip(), body.name.strip()
    if not code or not name:
        raise HTTPException(422, "employee_code and name are required")
    # Accepts the workbook's OWN/Supply as well as GI/NON_GI — the rename was
    # ours, and an integrator posting the old words should not get a 422.
    wt = normalize_worker_type(body.worker_type)
    if wt is None:
        raise HTTPException(422, f"worker_type must be one of "
                                 f"{list(MH_WORKER_TYPES)} (OWN / Supply are "
                                 f"accepted as the workbook's names for them)")
    shift = str(body.shift or "Day").strip().title()
    if shift not in MH_SHIFTS:
        raise HTTPException(422, f"shift must be one of {list(MH_SHIFTS)}")
    # Legend default (ADD EMPLOYEE sheet): GI→GI, NON_GI→DMC when blank.
    company = _clean(body.company) or _COMPANY_DEFAULTS[wt]
    stmt = pg_insert(employees_t).values(
        Site_ID=sid, Employee_Code=code, Name=name,
        Designation=_clean(body.designation) or "",
        Worker_Type=wt, Shift=shift, Company=company,
        status="active", created_by=user["username"])
    stmt = stmt.on_conflict_do_update(
        index_elements=["Site_ID", "Employee_Code"],
        set_={"Name": stmt.excluded.Name, "Designation": stmt.excluded.Designation,
              "Worker_Type": stmt.excluded.Worker_Type,
              "Shift": stmt.excluded.Shift, "Company": stmt.excluded.Company,
              "updated_at": _dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None)})
    await session.execute(stmt)
    await write_audit(session, user["username"], "MH_EMPLOYEE_UPSERT", "mh_employees",
                      f"{sid}/{code} {name}")
    await session.commit()
    return {"saved": True, "site_id": sid, "employee_code": code}


@router.patch("/employees/{emp_id}/status", summary="Flip a worker active/inactive")
async def set_employee_status(emp_id: int, status: str,
                              user: dict = Depends(require_roles("hod")),
                              session: AsyncSession = Depends(get_session)):
    if status not in ("active", "inactive"):
        raise HTTPException(422, "status must be active | inactive")
    stmt = update(employees_t).where(employees_t.c["id"] == emp_id).values(
        status=status, updated_at=_dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None))
    scope = site_scope(user)
    if scope is not None:
        stmt = stmt.where(employees_t.c["Site_ID"] == scope)
    res = await session.execute(stmt)
    if res.rowcount == 0:
        raise HTTPException(404, f"employee {emp_id} not found (or not your site)")
    await write_audit(session, user["username"], "MH_EMPLOYEE_STATUS", "mh_employees",
                      f"id={emp_id} → {status}")
    await session.commit()
    return {"id": emp_id, "status": status}


# --- Daily timesheets ------------------------------------------------------------
class TsRow(BaseModel):
    employee_code: str
    in_time: str
    out_time: str
    remarks: Optional[str] = ""


class TimesheetBatchIn(BaseModel):
    work_date: str            # YYYY-MM-DD
    equipment_tag: str
    system_code: str
    location: Optional[str] = ""
    break_mins: int = MH_DEFAULT_BREAK_MINS
    rows: list[TsRow]
    site_id: Optional[str] = None


async def _upsert_timesheet(session: AsyncSession, sid: str, code: str, wdate: str,
                            in_time, out_time, *, location: str = "",
                            equipment_tag: str = "", system_code: str = "",
                            break_mins: int = MH_DEFAULT_BREAK_MINS,
                            status: str = "PR", remarks: str = "",
                            created_by: str = "system",
                            threshold: Optional[float] = None) -> float:
    if threshold is None:
        # Single-row fallback. Batch callers pass `threshold` from
        # `thresholds_for_codes` instead — see its docstring for why.
        threshold = (await thresholds_for_codes(session, sid, [code])).get(
            code, (await ot_thresholds(session))["GI"])
    total, normal, ot = compute_mh_hours(in_time, out_time, break_mins,
                                         threshold)
    stmt = pg_insert(timesheets_t).values(
        Site_ID=sid, Employee_Code=code, Work_Date=str(wdate)[:10],
        Location=_clean(location),
        Equipment_Tag=_clean(equipment_tag),
        System_Code=_clean(system_code),
        In_Time="" if in_time is None else str(in_time),
        Out_Time="" if out_time is None else str(out_time),
        Break_Mins=int(break_mins or 0), Total_Hours=total, Normal_Hours=normal,
        OT_Hours=ot, Status=status or "PR", Remarks=(remarks or "").strip(),
        created_by=created_by)
    # NB: NULL Equipment_Tag/System_Code rows never conflict (PG treats NULLs as
    # distinct in the unique index) — same semantics as the legacy SQLite path;
    # the import's replace mode deletes-by-date first for exactly this reason.
    stmt = stmt.on_conflict_do_update(
        index_elements=["Site_ID", "Employee_Code", "Work_Date",
                        "Equipment_Tag", "System_Code"],
        set_={"In_Time": stmt.excluded.In_Time, "Out_Time": stmt.excluded.Out_Time,
              "Location": stmt.excluded.Location, "Break_Mins": stmt.excluded.Break_Mins,
              "Total_Hours": stmt.excluded.Total_Hours,
              "Normal_Hours": stmt.excluded.Normal_Hours,
              "OT_Hours": stmt.excluded.OT_Hours, "Status": stmt.excluded.Status,
              "Remarks": stmt.excluded.Remarks})
    await session.execute(stmt)
    return total


def _unassigned_cond():
    """Rows with no equipment linkage. Belt & braces: NULL is the canonical
    form, but ''/'nan' are matched too in case legacy tooling reintroduces
    them into SQLite → PG via dual_ci (the frozen legacy uploader can)."""
    tag = timesheets_t.c["Equipment_Tag"]
    return tag.is_(None) | func.lower(func.trim(tag)).in_(list(_BLANKISH - {""}) + [""])


@router.get("/timesheets", summary="Timesheet rows (flexible filters)")
async def list_timesheets(site_id: Optional[str] = None, work_date: Optional[str] = None,
                          employee_code: Optional[str] = None,
                          equipment_tag: Optional[str] = None,
                          date_from: Optional[str] = None, date_to: Optional[str] = None,
                          unassigned: bool = False,
                          user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    t = timesheets_t
    stmt = select(t.c["id"], t.c["Site_ID"], t.c["Employee_Code"], t.c["Work_Date"],
                  t.c["Location"], t.c["Equipment_Tag"], t.c["System_Code"],
                  t.c["In_Time"], t.c["Out_Time"], t.c["Break_Mins"],
                  t.c["Total_Hours"], t.c["Normal_Hours"], t.c["OT_Hours"],
                  t.c["Allocated_SQM"], t.c["Status"], t.c["Remarks"])
    if sid is not None:
        stmt = stmt.where(t.c["Site_ID"] == sid)
    for col, val in (("Work_Date", work_date), ("Employee_Code", employee_code),
                     ("Equipment_Tag", equipment_tag)):
        if val:
            stmt = stmt.where(t.c[col] == val)
    if date_from:
        stmt = stmt.where(t.c["Work_Date"] >= date_from)
    if date_to:
        stmt = stmt.where(t.c["Work_Date"] <= date_to)
    if unassigned:
        stmt = stmt.where(_unassigned_cond())
    stmt = stmt.order_by(t.c["Work_Date"].desc(), t.c["Employee_Code"]).limit(1000)
    items = _rows(await session.execute(stmt))
    return {"items": items,
            "total_hours": round(sum(float(r["Total_Hours"] or 0) for r in items), 1)}


@router.post("/timesheets", summary="Save a per-day batch of timesheet rows")
async def save_timesheets(body: TimesheetBatchIn = Body(...),
                          user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    sid = _write_site(user, body.site_id)
    if not body.work_date.strip() or not body.equipment_tag.strip() \
            or not body.system_code.strip():
        raise HTTPException(422, "work_date, equipment_tag and system_code are required")
    if not body.rows:
        raise HTTPException(422, "no rows to save")
    saved = 0
    thr = await thresholds_for_codes(session, sid,
                                     [r.employee_code for r in body.rows])
    gi = (await ot_thresholds(session))["GI"]
    for r in body.rows:
        if not r.employee_code.strip():
            continue
        await _upsert_timesheet(
            session, sid, r.employee_code.strip(), body.work_date,
            r.in_time, r.out_time, location=body.location or "",
            equipment_tag=body.equipment_tag, system_code=body.system_code,
            break_mins=body.break_mins, remarks=r.remarks or "",
            created_by=user["username"],
            threshold=thr.get(r.employee_code.strip(), gi))
        saved += 1
    await write_audit(session, user["username"], "MH_TIMESHEET_BATCH", "mh_timesheets",
                      f"{sid} {body.work_date} {body.equipment_tag}/{body.system_code} "
                      f"rows={saved}")
    await session.commit()
    return {"saved": saved, "work_date": body.work_date}


@router.delete("/timesheets/{ts_id}", summary="Delete one timesheet row")
async def delete_timesheet(ts_id: int, user: dict = Depends(require_roles("hod")),
                           session: AsyncSession = Depends(get_session)):
    stmt = delete(timesheets_t).where(timesheets_t.c["id"] == ts_id)
    scope = site_scope(user)
    if scope is not None:
        stmt = stmt.where(timesheets_t.c["Site_ID"] == scope)
    res = await session.execute(stmt)
    if res.rowcount == 0:
        raise HTTPException(404, f"timesheet {ts_id} not found (or not your site)")
    await write_audit(session, user["username"], "MH_TIMESHEET_DELETE", "mh_timesheets",
                      f"id={ts_id}")
    await session.commit()
    return {"deleted": ts_id}


# --- Bulk-assign: tie unassigned hours to an SME scope -----------------------------
class AssignIn(BaseModel):
    ids: list[int]
    equipment_tag: str
    system_code: str
    location: Optional[str] = None  # blank → auto-filled from sme_equipment
    site_id: Optional[str] = None


@router.patch("/timesheets/assign",
              summary="Assign timesheet rows to an Equipment/System scope (bulk)")
async def assign_timesheets(body: AssignIn = Body(...),
                            user: dict = Depends(require_roles("hod")),
                            session: AsyncSession = Depends(get_session)):
    """The attendance workbook ships with Equipment Tag # blank, so imported
    hours land unassigned. This is the workflow that ties them to a scope so
    they count in Estimate-vs-Actual (and, later, the SME scorecard).

    Rows whose (site, employee, date) already have a row on the TARGET scope
    would collide with the unique key — those are skipped and reported, never
    merged silently."""
    sid = _write_site(user, body.site_id)
    tag, sc = _clean(body.equipment_tag), _clean(body.system_code)
    if not tag or not sc:
        raise HTTPException(422, "equipment_tag and system_code are required")
    if not body.ids:
        raise HTTPException(422, "no timesheet ids given")
    if len(body.ids) > 500:
        raise HTTPException(422, "at most 500 rows per assign call")

    location = _clean(body.location)
    if location is None:
        e = sme_equipment_t
        location = (await session.execute(
            select(e.c["Location"]).where(e.c["Equipment_Tag_No"] == tag)
            .order_by(e.c["id"]).limit(1))).scalar()

    t = timesheets_t
    rows = _rows(await session.execute(
        select(t.c["id"], t.c["Employee_Code"], t.c["Work_Date"],
               t.c["Equipment_Tag"], t.c["System_Code"])
        .where(t.c["id"].in_(body.ids), t.c["Site_ID"] == sid)))
    found = {r["id"] for r in rows}
    missing = [i for i in body.ids if i not in found]

    assigned, conflicts = [], []
    for r in rows:
        if r["Equipment_Tag"] == tag and r["System_Code"] == sc:
            continue  # already on the target scope — nothing to do
        dup = (await session.execute(select(func.count()).select_from(t).where(
            t.c["Site_ID"] == sid, t.c["Employee_Code"] == r["Employee_Code"],
            t.c["Work_Date"] == r["Work_Date"], t.c["Equipment_Tag"] == tag,
            t.c["System_Code"] == sc, t.c["id"] != r["id"]))).scalar_one()
        if dup:
            conflicts.append({"id": r["id"], "employee_code": r["Employee_Code"],
                              "work_date": r["Work_Date"],
                              "reason": "a row for this worker/date already exists on the target scope"})
            continue
        await session.execute(update(t).where(t.c["id"] == r["id"]).values(
            Equipment_Tag=tag, System_Code=sc, Location=location))
        assigned.append(r["id"])

    await write_audit(session, user["username"], "MH_TIMESHEET_ASSIGN", "mh_timesheets",
                      f"{sid} → {tag}/{sc} assigned={len(assigned)} "
                      f"conflicts={len(conflicts)} missing={len(missing)}")
    await session.commit()
    return {"assigned": len(assigned), "ids": assigned, "conflicts": conflicts,
            "missing": missing, "equipment_tag": tag, "system_code": sc,
            "location": location}


# --- Team SQM production + distribution -------------------------------------------
class ProductionIn(BaseModel):
    work_date: str
    equipment_tag: str
    system_code: str
    sqm_done: float
    distribution_method: str = "even"  # even | by_hours
    site_id: Optional[str] = None


@router.post("/production", summary="Record team SQM and distribute it to workers")
async def set_production(body: ProductionIn = Body(...),
                         user: dict = Depends(require_roles("hod")),
                         session: AsyncSession = Depends(get_session)):
    if body.distribution_method not in ("even", "by_hours"):
        raise HTTPException(422, "distribution_method must be even | by_hours")
    sid = _write_site(user, body.site_id)
    wdate = body.work_date.strip()[:10]
    stmt = pg_insert(production_t).values(
        Site_ID=sid, Work_Date=wdate, Equipment_Tag=body.equipment_tag.strip(),
        System_Code=body.system_code.strip(), SQM_Done=float(body.sqm_done or 0),
        Distribution_Method=body.distribution_method, created_by=user["username"])
    stmt = stmt.on_conflict_do_update(
        index_elements=["Site_ID", "Work_Date", "Equipment_Tag", "System_Code"],
        set_={"SQM_Done": stmt.excluded.SQM_Done,
              "Distribution_Method": stmt.excluded.Distribution_Method})
    await session.execute(stmt)

    # Distribute into that day's Allocated_SQM: even split, or pro-rata on hours.
    t = timesheets_t
    rows = (await session.execute(select(t.c["id"], t.c["Total_Hours"]).where(
        t.c["Site_ID"] == sid, t.c["Work_Date"] == wdate,
        t.c["Equipment_Tag"] == body.equipment_tag.strip(),
        t.c["System_Code"] == body.system_code.strip()))).all()
    total_sqm = float(body.sqm_done or 0)
    updated = 0
    if rows:
        if body.distribution_method == "by_hours":
            hours_sum = sum(float(h or 0) for _, h in rows)
            for rid, hrs in rows:
                share = (total_sqm * float(hrs or 0) / hours_sum) if hours_sum else 0.0
                await session.execute(update(t).where(t.c["id"] == rid)
                                      .values(Allocated_SQM=round(share, 3)))
                updated += 1
        else:
            share = total_sqm / len(rows)
            for rid, _hrs in rows:
                await session.execute(update(t).where(t.c["id"] == rid)
                                      .values(Allocated_SQM=round(share, 3)))
                updated += 1
    await write_audit(session, user["username"], "MH_PRODUCTION_SET", "mh_production",
                      f"{sid} {wdate} {body.equipment_tag}/{body.system_code} "
                      f"sqm={total_sqm:g} {body.distribution_method} rows={updated}")
    await session.commit()
    return {"saved": True, "distributed_rows": updated}


# --- Man-hour estimator -------------------------------------------------------------
class EstimateIn(BaseModel):
    equipment_tag: str
    system_code: str
    estimated_manhours: float
    estimated_sqm: Optional[float] = None
    location: Optional[str] = ""
    basis: Optional[str] = ""
    site_id: Optional[str] = None


@router.get("/estimates", summary="Required man-hours per Tag/System")
async def list_estimates(site_id: Optional[str] = None,
                         user: dict = Depends(require_roles("hod")),
                         session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    t = estimates_t
    stmt = select(t.c["id"], t.c["Site_ID"], t.c["Location"], t.c["Equipment_Tag"],
                  t.c["System_Code"], t.c["Estimated_Manhours"], t.c["Estimated_SQM"],
                  t.c["Basis"], t.c["created_at"])
    if sid is not None:
        stmt = stmt.where(t.c["Site_ID"] == sid)
    stmt = stmt.order_by(t.c["Equipment_Tag"], t.c["System_Code"])
    return {"items": _rows(await session.execute(stmt))}


@router.post("/estimates", summary="Define/update an estimate (upsert on Tag+System)")
async def upsert_estimate(body: EstimateIn = Body(...),
                          user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    sid = _write_site(user, body.site_id)
    tag, sc = body.equipment_tag.strip(), body.system_code.strip()
    if not tag or not sc:
        raise HTTPException(422, "equipment_tag and system_code are required")
    if body.estimated_manhours < 0:
        raise HTTPException(422, "estimated_manhours must be ≥ 0")
    stmt = pg_insert(estimates_t).values(
        Site_ID=sid, Location=(body.location or "").strip() or None,
        Equipment_Tag=tag, System_Code=sc,
        Estimated_Manhours=float(body.estimated_manhours),
        Estimated_SQM=body.estimated_sqm, Basis=(body.basis or "").strip(),
        created_by=user["username"])
    stmt = stmt.on_conflict_do_update(
        index_elements=["Site_ID", "Equipment_Tag", "System_Code"],
        set_={"Location": stmt.excluded.Location,
              "Estimated_Manhours": stmt.excluded.Estimated_Manhours,
              "Estimated_SQM": stmt.excluded.Estimated_SQM,
              "Basis": stmt.excluded.Basis,
              "updated_at": _dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None)})
    await session.execute(stmt)
    await write_audit(session, user["username"], "MH_ESTIMATE_UPSERT",
                      "mh_manhour_estimates",
                      f"{sid} {tag}/{sc} mh={body.estimated_manhours:g}")
    await session.commit()
    return {"saved": True, "equipment_tag": tag, "system_code": sc}


@router.delete("/estimates/{est_id}", summary="Remove an estimate")
async def delete_estimate(est_id: int, user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    stmt = delete(estimates_t).where(estimates_t.c["id"] == est_id)
    scope = site_scope(user)
    if scope is not None:
        stmt = stmt.where(estimates_t.c["Site_ID"] == scope)
    res = await session.execute(stmt)
    if res.rowcount == 0:
        raise HTTPException(404, f"estimate {est_id} not found (or not your site)")
    await write_audit(session, user["username"], "MH_ESTIMATE_DELETE",
                      "mh_manhour_estimates", f"id={est_id}")
    await session.commit()
    return {"deleted": est_id}


# --- Phase 11C: planning automation over the 11B join layer ------------------------
async def _auto_draft_items(session: AsyncSession, sid: Optional[str],
                            norm_override: Optional[float]) -> dict:
    """Draft labor estimates for every SME scope that has remaining SQM and no
    estimate yet: remaining × MH/SQM norm. The scope's own learned norm wins;
    otherwise the site norm (or an explicit override). READ-ONLY against sme_*."""
    sme = await _sme_scopes(session, sid)
    prog = await _sme_progress(session, sid)
    est = await _estimate_map(session, sid)
    prodv = await _productivity_rows(session, sid)
    scope_norms = {(r["Equipment_Tag"], r["System_Code"]): r["MH_per_SQM"]
                   for r in prodv["items"] if r["MH_per_SQM"]}
    site_norm = norm_override if norm_override else prodv["site_norm"]["mh_per_sqm"]

    items = []
    for key in sorted(sme, key=lambda k: (str(k[0]), str(k[1]))):
        if key in est:
            continue  # already estimated — never overwrite silently
        s, p = sme[key], prog.get(key)
        planned = float((p and p["original_sqm"]) or s["surface_sqm"] or 0)
        done = float((p and p["done_sqm"]) or 0)
        remaining = max(planned - done, 0.0)
        if remaining <= 0:
            continue
        if norm_override:
            norm, source = norm_override, "override"
        elif key in scope_norms:
            norm, source = scope_norms[key], "scope"
        else:
            norm, source = site_norm, "site"
        items.append({
            "Equipment_Tag": key[0], "System_Code": key[1],
            "Location": s["location"], "Remaining_SQM": round(remaining, 2),
            "Norm_Used": round(norm, 3) if norm else None,
            "Norm_Source": source if norm else None,
            "Draft_Manhours": round(remaining * norm, 1) if norm else None,
        })
    return {"items": items, "site_norm": site_norm,
            "hint": None if site_norm else
            "no productivity history yet — pass ?norm= to draft with a manual MH/SQM norm"}


@router.get("/estimates/auto-draft",
            summary="Preview draft estimates: SME remaining SQM × MH/SQM norm")
async def auto_draft_preview(site_id: Optional[str] = None, norm: Optional[float] = None,
                             user: dict = Depends(require_roles("hod")),
                             session: AsyncSession = Depends(get_session)):
    if norm is not None and norm <= 0:
        raise HTTPException(422, "norm must be > 0")
    sid = resolve_site_param(user, site_id)
    return await _auto_draft_items(session, sid, norm)


class DraftRow(BaseModel):
    equipment_tag: str
    system_code: str
    estimated_manhours: float
    estimated_sqm: Optional[float] = None
    location: Optional[str] = None
    basis: Optional[str] = None


class AutoDraftIn(BaseModel):
    rows: list[DraftRow]
    site_id: Optional[str] = None


@router.post("/estimates/auto-draft",
             summary="Save reviewed draft estimates (bulk upsert into mh_manhour_estimates)")
async def auto_draft_save(body: AutoDraftIn = Body(...),
                          user: dict = Depends(require_roles("hod")),
                          session: AsyncSession = Depends(get_session)):
    sid = _write_site(user, body.site_id)
    if not body.rows:
        raise HTTPException(422, "no rows to save")
    if len(body.rows) > 200:
        raise HTTPException(422, "at most 200 estimates per save")
    saved = 0
    for row in body.rows:
        tag, sc = _clean(row.equipment_tag), _clean(row.system_code)
        if not tag or not sc or row.estimated_manhours < 0:
            raise HTTPException(422, f"bad draft row: {row.equipment_tag}/{row.system_code}")
        stmt = pg_insert(estimates_t).values(
            Site_ID=sid, Location=_clean(row.location), Equipment_Tag=tag,
            System_Code=sc, Estimated_Manhours=float(row.estimated_manhours),
            Estimated_SQM=row.estimated_sqm,
            Basis=_clean(row.basis) or "auto-draft (SME remaining SQM × norm)",
            created_by=user["username"])
        stmt = stmt.on_conflict_do_update(
            index_elements=["Site_ID", "Equipment_Tag", "System_Code"],
            set_={"Location": stmt.excluded.Location,
                  "Estimated_Manhours": stmt.excluded.Estimated_Manhours,
                  "Estimated_SQM": stmt.excluded.Estimated_SQM,
                  "Basis": stmt.excluded.Basis,
                  "updated_at": _dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None)})
        await session.execute(stmt)
        saved += 1
    await write_audit(session, user["username"], "MH_ESTIMATE_AUTODRAFT",
                      "mh_manhour_estimates", f"{sid} rows={saved}")
    await session.commit()
    return {"saved": saved}


@router.get("/forecast",
            summary="Manpower forecast: days-to-complete per scope for a crew size")
async def forecast(crew_size: int = 10, hours_per_day: float = 8.0,
                   site_id: Optional[str] = None,
                   user: dict = Depends(require_roles("hod")),
                   session: AsyncSession = Depends(get_session)):
    """Remaining man-hours per scope ÷ (crew × hours/day). Scopes WITH an
    estimate use max(estimated − actual, 0); scopes without one fall back to
    remaining SQM × the productivity norm. Fully-consumed scopes drop out."""
    if not (1 <= crew_size <= 1000):
        raise HTTPException(422, "crew_size must be 1–1000")
    if not (1.0 <= hours_per_day <= 24.0):
        raise HTTPException(422, "hours_per_day must be 1–24")
    sid = resolve_site_param(user, site_id)

    est = await _estimate_map(session, sid)
    hours = await _labor_hours(session, sid)
    drafts = await _auto_draft_items(session, sid, None)  # unestimated scopes
    capacity = crew_size * hours_per_day

    items = []
    for key, e in sorted(est.items(), key=lambda kv: (str(kv[0][0]), str(kv[0][1]))):
        actual = float(hours.get(key, {}).get("hours") or 0)
        remaining_mh = max(float(e["est_mh"]) - actual, 0.0)
        if remaining_mh <= 0:
            continue
        items.append({"Equipment_Tag": key[0], "System_Code": key[1],
                      "Basis": "estimate", "Remaining_SQM": None,
                      "Remaining_Manhours": round(remaining_mh, 1),
                      "Days_To_Complete": round(remaining_mh / capacity, 2)})
    for d in drafts["items"]:
        if d["Draft_Manhours"] is None:
            continue
        items.append({"Equipment_Tag": d["Equipment_Tag"],
                      "System_Code": d["System_Code"], "Basis": "norm",
                      "Remaining_SQM": d["Remaining_SQM"],
                      "Remaining_Manhours": d["Draft_Manhours"],
                      "Days_To_Complete": round(d["Draft_Manhours"] / capacity, 2)})
    items.sort(key=lambda r: (str(r["Equipment_Tag"]), str(r["System_Code"])))
    total_mh = round(sum(r["Remaining_Manhours"] for r in items), 1)
    return {"items": items, "crew_size": crew_size, "hours_per_day": hours_per_day,
            "rollup": {"scopes": len(items), "total_remaining_manhours": total_mh,
                       "days_to_complete": round(total_mh / capacity, 1),
                       "site_norm": drafts["site_norm"]}}


# --- Estimate vs Actual (inline port of the legacy v_mh_estimate_vs_actual view) ---
SQL_MH_VARIANCE = '''
SELECT e."Site_ID", e."Equipment_Tag", e."System_Code", e."Location",
       e."Estimated_Manhours",
       COALESCE(a.actual, 0)                          AS "Actual_Manhours",
       COALESCE(a.actual, 0) - e."Estimated_Manhours" AS "Variance_Manhours",
       CASE WHEN e."Estimated_Manhours" > 0
            THEN ROUND(CAST((COALESCE(a.actual, 0) - e."Estimated_Manhours") * 100.0
                            / e."Estimated_Manhours" AS NUMERIC), 1)
            ELSE NULL END                             AS "Variance_Pct",
       COALESCE(p.sqm, 0)                             AS "SQM_Done",
       n."Reason"                                     AS "Variance_Reason"
FROM mh_manhour_estimates e
LEFT JOIN (SELECT "Site_ID", "Equipment_Tag", "System_Code",
                  SUM("Total_Hours") AS actual
           FROM mh_timesheets GROUP BY 1, 2, 3) a
       ON a."Site_ID" = e."Site_ID" AND a."Equipment_Tag" = e."Equipment_Tag"
      AND a."System_Code" = e."System_Code"
LEFT JOIN (SELECT "Site_ID", "Equipment_Tag", "System_Code",
                  SUM("SQM_Done") AS sqm
           FROM mh_production GROUP BY 1, 2, 3) p
       ON p."Site_ID" = e."Site_ID" AND p."Equipment_Tag" = e."Equipment_Tag"
      AND p."System_Code" = e."System_Code"
LEFT JOIN mh_variance_notes n
       ON n."Site_ID" = e."Site_ID" AND n."Equipment_Tag" = e."Equipment_Tag"
      AND n."System_Code" = e."System_Code"
{where}
ORDER BY "Variance_Manhours" DESC'''


async def _variance_rows(session: AsyncSession, site_id: Optional[str]) -> list[dict]:
    where, params = "", {}
    if site_id is not None:
        where, params = 'WHERE e."Site_ID" = :site', {"site": site_id}
    return _rows(await session.execute(text(SQL_MH_VARIANCE.format(where=where)), params))


@router.get("/variance", summary="Estimate-vs-Actual dashboard rows")
async def variance(site_id: Optional[str] = None,
                   user: dict = Depends(require_roles("hod")),
                   session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    items = await _variance_rows(session, sid)
    over = [r for r in items if float(r["Variance_Manhours"] or 0) > 0]
    return {"items": items,
            "kpis": {"scopes": len(items), "over_consuming": len(over),
                     "total_actual": round(sum(float(r["Actual_Manhours"] or 0)
                                               for r in items), 1)}}


class ReasonIn(BaseModel):
    equipment_tag: str
    system_code: str
    reason: str
    site_id: Optional[str] = None


@router.post("/variance/reason", summary="Record an over-consumption reason")
async def set_variance_reason(body: ReasonIn = Body(...),
                              user: dict = Depends(require_roles("hod")),
                              session: AsyncSession = Depends(get_session)):
    if not body.reason.strip():
        raise HTTPException(422, "reason is required")
    sid = _write_site(user, body.site_id)
    stmt = pg_insert(variance_t).values(
        Site_ID=sid, Equipment_Tag=body.equipment_tag.strip(),
        System_Code=body.system_code.strip(), Reason=body.reason.strip(),
        entered_by=user["username"])
    stmt = stmt.on_conflict_do_update(
        index_elements=["Site_ID", "Equipment_Tag", "System_Code"],
        set_={"Reason": stmt.excluded.Reason, "entered_by": stmt.excluded.entered_by,
              "created_at": _dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None)})
    await session.execute(stmt)
    await write_audit(session, user["username"], "MH_VARIANCE_REASON",
                      "mh_variance_notes",
                      f"{sid} {body.equipment_tag}/{body.system_code}")
    await session.commit()
    return {"saved": True}


# --- Employee-wise timeline ---------------------------------------------------------
@router.get("/employee-timeline", summary="Where each worker worked, date by date")
async def employee_timeline(site_id: Optional[str] = None,
                            employee_code: Optional[str] = None,
                            date_from: Optional[str] = None,
                            date_to: Optional[str] = None,
                            user: dict = Depends(require_roles("hod")),
                            session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    t, e = timesheets_t, employees_t
    stmt = (select(t.c["Employee_Code"],
                   e.c["Name"].label("Name"),
                   t.c["Work_Date"], t.c["Location"], t.c["Equipment_Tag"],
                   t.c["System_Code"], t.c["Total_Hours"], t.c["Normal_Hours"],
                   t.c["OT_Hours"], t.c["Allocated_SQM"])
            .join(e, (e.c["Site_ID"] == t.c["Site_ID"])
                  & (e.c["Employee_Code"] == t.c["Employee_Code"]), isouter=True))
    if sid is not None:
        stmt = stmt.where(t.c["Site_ID"] == sid)
    if employee_code:
        stmt = stmt.where(t.c["Employee_Code"] == employee_code)
    if date_from:
        stmt = stmt.where(t.c["Work_Date"] >= date_from)
    if date_to:
        stmt = stmt.where(t.c["Work_Date"] <= date_to)
    stmt = stmt.order_by(t.c["Employee_Code"], t.c["Work_Date"]).limit(2000)
    items = _rows(await session.execute(stmt))
    for r in items:
        r["Name"] = r["Name"] or r["Employee_Code"]
    return {"items": items,
            "total_hours": round(sum(float(r["Total_Hours"] or 0) for r in items), 1)}


# --- Phase 11B: SME ↔ MH link layer (READ-ONLY joins — SME Canon) -------------------
# Both domains share the natural key (Site_ID, Equipment_Tag == Equipment_Tag_No,
# System_Code == Lining_System_Code). Everything below is SELECT-only against
# sme_* ; the merge happens in Python (≤ ~100 scopes — same style as sme.py).

async def _scope_map(session: AsyncSession, stmt, keys=("tag", "sys")) -> dict:
    """Execute a (tag, sys, *values) grouped SELECT → {(tag, sys): row-dict}."""
    out = {}
    for m in (await session.execute(stmt)).mappings().all():
        d = dict(m)
        out[(d.pop(keys[0]), d.pop(keys[1]))] = d
    return out


async def _labor_hours(session: AsyncSession, sid: Optional[str]) -> dict:
    t = timesheets_t
    stmt = (select(t.c["Equipment_Tag"].label("tag"), t.c["System_Code"].label("sys"),
                   func.sum(t.c["Total_Hours"]).label("hours"))
            .where(t.c["Equipment_Tag"].is_not(None), t.c["System_Code"].is_not(None))
            .group_by(t.c["Equipment_Tag"], t.c["System_Code"]))
    if sid is not None:
        stmt = stmt.where(t.c["Site_ID"] == sid)
    return await _scope_map(session, stmt)


async def _labor_sqm(session: AsyncSession, sid: Optional[str]) -> dict:
    p = production_t
    stmt = (select(p.c["Equipment_Tag"].label("tag"), p.c["System_Code"].label("sys"),
                   func.sum(p.c["SQM_Done"]).label("sqm"))
            .group_by(p.c["Equipment_Tag"], p.c["System_Code"]))
    if sid is not None:
        stmt = stmt.where(p.c["Site_ID"] == sid)
    return await _scope_map(session, stmt)


async def _estimate_map(session: AsyncSession, sid: Optional[str]) -> dict:
    e = estimates_t
    stmt = select(e.c["Equipment_Tag"].label("tag"), e.c["System_Code"].label("sys"),
                  e.c["Estimated_Manhours"].label("est_mh"),
                  e.c["Estimated_SQM"].label("est_sqm"))
    if sid is not None:
        stmt = stmt.where(e.c["Site_ID"] == sid)
    return await _scope_map(session, stmt)


async def _sme_scopes(session: AsyncSession, sid: Optional[str]) -> dict:
    """(tag, system) → location + planned surface from sme_equipment (READ-ONLY).
    Area rows can repeat per scope → SUM the surface, keep the first location."""
    e = sme_equipment_t
    stmt = (select(e.c["Equipment_Tag_No"].label("tag"),
                   e.c["Lining_System_Code"].label("sys"),
                   func.min(e.c["Location"]).label("location"),
                   func.sum(e.c["Surface_Area_SQM"]).label("surface_sqm"))
            .where(e.c["Equipment_Tag_No"].is_not(None))
            .group_by(e.c["Equipment_Tag_No"], e.c["Lining_System_Code"]))
    if sid is not None:
        stmt = stmt.where(e.c["Site_ID"] == sid)
    return await _scope_map(session, stmt)


async def _sme_progress(session: AsyncSession, sid: Optional[str]) -> dict:
    s = _MD.tables["sme_sqm_progress"]
    stmt = (select(s.c["Equipment_Tag_No"].label("tag"),
                   s.c["Lining_System_Code"].label("sys"),
                   func.sum(s.c["Original_SQM"]).label("original_sqm"),
                   func.sum(s.c["Done_SQM"]).label("done_sqm"))
            .group_by(s.c["Equipment_Tag_No"], s.c["Lining_System_Code"]))
    if sid is not None:
        stmt = stmt.where(s.c["Site_ID"] == sid)
    return await _scope_map(session, stmt)


async def _material_variance(session: AsyncSession, sid: Optional[str]) -> dict:
    """(tag, system) → expected vs actual material qty from sme_consumption_log
    (READ-ONLY; rejected entries excluded)."""
    c = _MD.tables["sme_consumption_log"]
    stmt = (select(c.c["Equipment_Tag_No"].label("tag"),
                   c.c["Lining_System_Code"].label("sys"),
                   func.sum(c.c["Expected_Qty"]).label("mat_expected"),
                   func.sum(c.c["Actual_Qty"]).label("mat_actual"))
            .where(c.c["status"] != "rejected")
            .group_by(c.c["Equipment_Tag_No"], c.c["Lining_System_Code"]))
    if sid is not None:
        stmt = stmt.where(c.c["Site_ID"] == sid)
    return await _scope_map(session, stmt)


def _pct(actual: float, base: float) -> Optional[float]:
    return round((actual - base) * 100.0 / base, 1) if base else None


def _recon(done_labor: float, done_sme: float) -> Optional[str]:
    """Two independent 'SQM done' sources (labor-reported vs SME-reported).
    None = nothing measured yet; 'drift' when they disagree by > max(1, 5%)."""
    top = max(done_labor, done_sme)
    if top <= 0:
        return None
    return "drift" if abs(done_labor - done_sme) > max(1.0, 0.05 * top) else "ok"


async def _productivity_rows(session: AsyncSession, sid: Optional[str]) -> dict:
    hours = await _labor_hours(session, sid)
    sqm = await _labor_sqm(session, sid)
    est = await _estimate_map(session, sid)
    items = []
    for key in sorted(set(hours) | set(sqm), key=lambda k: (str(k[0]), str(k[1]))):
        h = float(hours.get(key, {}).get("hours") or 0)
        q = float(sqm.get(key, {}).get("sqm") or 0)
        e = est.get(key, {})
        est_norm = None
        if e.get("est_sqm") and float(e["est_sqm"]) > 0:
            est_norm = round(float(e["est_mh"]) / float(e["est_sqm"]), 3)
        items.append({
            "Equipment_Tag": key[0], "System_Code": key[1],
            "Actual_Manhours": round(h, 1), "SQM_Done": round(q, 2),
            "MH_per_SQM": round(h / q, 3) if q > 0 else None,
            "SQM_per_MH": round(q / h, 3) if h > 0 else None,
            "Est_MH_per_SQM": est_norm,
        })
    th = sum(r["Actual_Manhours"] for r in items if r["SQM_Done"] > 0)
    tq = sum(r["SQM_Done"] for r in items if r["Actual_Manhours"] > 0)
    site_norm = {
        "hours": round(th, 1), "sqm": round(tq, 2),
        "mh_per_sqm": round(th / tq, 3) if tq > 0 else None,
        "sqm_per_mh": round(tq / th, 3) if th > 0 else None,
    }
    return {"items": items, "site_norm": site_norm}


@router.get("/productivity", summary="Labor norms per scope + the site norm (MH/SQM)")
async def productivity(site_id: Optional[str] = None,
                       user: dict = Depends(require_roles("hod")),
                       session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    return await _productivity_rows(session, sid)


async def _scorecard_rows(session: AsyncSession, sid: Optional[str]) -> dict:
    sme = await _sme_scopes(session, sid)
    prog = await _sme_progress(session, sid)
    mat = await _material_variance(session, sid)
    hours = await _labor_hours(session, sid)
    lsqm = await _labor_sqm(session, sid)
    est = await _estimate_map(session, sid)

    keys = set(sme) | set(hours) | set(lsqm) | set(est)
    items = []
    for key in sorted(keys, key=lambda k: (str(k[0]), str(k[1]))):
        s, p, m = sme.get(key), prog.get(key), mat.get(key)
        h = float(hours.get(key, {}).get("hours") or 0)
        dl = float(lsqm.get(key, {}).get("sqm") or 0)
        e = est.get(key, {})
        planned = float((p and p["original_sqm"]) or (s and s["surface_sqm"]) or 0)
        done_sme = float((p and p["done_sqm"]) or 0)
        est_mh = None if not e else float(e["est_mh"])
        mat_exp = float((m and m["mat_expected"]) or 0)
        mat_act = float((m and m["mat_actual"]) or 0)
        items.append({
            "Equipment_Tag": key[0], "System_Code": key[1],
            "Location": s["location"] if s else None,
            "In_SME": s is not None,
            "Planned_SQM": round(planned, 2) or None,
            "Done_SQM_SME": round(done_sme, 2),
            "Done_SQM_Labor": round(dl, 2),
            "Pct_Complete": round(100 * done_sme / planned, 1) if planned else None,
            "Estimated_Manhours": est_mh,
            "Actual_Manhours": round(h, 1),
            "Labor_Variance_Pct": _pct(h, est_mh) if est_mh else None,
            "MH_per_SQM": round(h / dl, 3) if dl > 0 else None,
            "Material_Expected": round(mat_exp, 2) or None,
            "Material_Actual": round(mat_act, 2) or None,
            "Material_Variance_Pct": _pct(mat_act, mat_exp),
            "Reconciliation": _recon(dl, done_sme),
        })
    kpis = {
        "scopes": len(items),
        "with_labor": sum(1 for r in items if r["Actual_Manhours"] > 0),
        "with_estimate": sum(1 for r in items if r["Estimated_Manhours"]),
        "drift": sum(1 for r in items if r["Reconciliation"] == "drift"),
        "total_hours": round(sum(r["Actual_Manhours"] for r in items), 1),
    }
    return {"items": items, "kpis": kpis,
            "site_norm": (await _productivity_rows(session, sid))["site_norm"]}


@router.get("/scorecard",
            summary="Unified per-equipment view: SME SQM + material vs labor variance")
async def scorecard(site_id: Optional[str] = None,
                    user: dict = Depends(require_roles("hod")),
                    session: AsyncSession = Depends(get_session)):
    sid = resolve_site_param(user, site_id)
    return await _scorecard_rows(session, sid)


# --- Attendance workbook import (openpyxl port of parse_attendance_workbook) --------
def _norm(s) -> str:
    return str(s or "").strip().lower()


def _str_code(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip()


def _iso_date(v) -> str:
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    return str(v or "").strip()[:10]


def _sheet_rows(ws) -> list[dict]:
    """First row = header; remaining rows keyed by normalized header name."""
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return []
    keys = [_norm(h) for h in header]
    return [dict(zip(keys, row)) for row in it]


def parse_attendance_workbook(data: bytes) -> dict:
    """Pure parse of the to-john_Attendance .xlsx format (no DB writes).
    ADD EMPLOYEE sheet supplies richer attributes; every distinct SAR worker
    is merged into the roster. Hours are recomputed downstream from In/Out."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    emp_rows: list[dict] = []
    if "ADD EMPLOYEE" in wb.sheetnames:
        for r in _sheet_rows(wb["ADD EMPLOYEE"]):
            code = _str_code(r.get("code"))
            name = str(r.get("name") or "").strip()
            if not code or not name:
                continue
            wt = normalize_worker_type(r.get("type")) or "GI"
            emp_rows.append({
                "code": code, "name": name,
                "designation": _clean(r.get("designation")) or "",
                "worker_type": wt,
                # Legend default: OWN→GI, Supply→DMC when the cell is blank.
                "company": _clean(r.get("company")) or _COMPANY_DEFAULTS[wt]})

    # In-file dedupe on the upsert key (code, date, tag): the last occurrence
    # wins, mirroring what the row-by-row upsert would have produced anyway —
    # but this also protects NULL-tag rows, which never conflict in PG.
    by_key: dict[tuple, dict] = {}
    if "SAR" in wb.sheetnames:
        for r in _sheet_rows(wb["SAR"]):
            code = _str_code(r.get("code"))
            wdate = _iso_date(r.get("work date"))
            if not code or not wdate:
                continue
            row = {
                "code": code, "name": str(r.get("name") or "").strip(),
                "work_date": wdate,
                "location": _clean(r.get("location")) or "",
                "equipment_tag": _clean(r.get("equipment tag #")
                                        or r.get("equipment tag")) or "",
                "in_time": r.get("in time"), "out_time": r.get("out time"),
                "status": str(r.get("status") or "").strip() or "PR",
                "remarks": _clean(r.get("remarks")) or ""}
            by_key[(code, wdate, row["equipment_tag"])] = row
    timesheets = list(by_key.values())

    by_code = {e["code"]: e for e in emp_rows}
    for t in timesheets:
        by_code.setdefault(t["code"], {
            "code": t["code"], "name": t["name"] or t["code"],
            "designation": "", "worker_type": "GI",
            "company": _COMPANY_DEFAULTS["GI"]})
    dates = sorted({t["work_date"] for t in timesheets})
    return {"employees": list(by_code.values()), "timesheets": timesheets,
            "dates": dates}


async def _sync_employee_master(session: AsyncSession, site_id: str,
                                rows: list[dict]) -> int:
    """QSEP — mirror the roster into `employees`, the PERSON registry.

    This import is the only bulk employee upload in the product, and before
    QSEP it wrote `mh_employees` ONLY. So a worker imported from the
    attendance workbook could not be named on a supervisor material request:
    `create_smr` looks them up in `employees` and answers "worker not in
    employee master". Two registries, no join, and the join column
    (`linked_id_number`) had never been written by anything.

    ⚠️ `employees.ID_Number` is UNIQUE **globally**, not per site. The same
    person working at two sites is ONE row, so this must UPDATE the site
    rather than insert a second — a naive insert 409s on the second site's
    import and the whole transaction fails.

    Name/designation/company are refreshed from the workbook; `Site_ID` is
    moved to the importing site, because importing somebody's attendance at
    a site IS the statement that they work there. Deliberately NOT touched:
    `status` (an employee deactivated in the app stays deactivated — the
    workbook has no status column and would silently revive them) and
    `Phone_Number` (entered in the app, absent from the workbook).
    """
    n = 0
    for e in rows:
        code = str(e["code"]).strip()
        if not code:
            continue
        stmt = pg_insert(employees_t_master).values(
            ID_Number=code, Name=e["name"], Site_ID=site_id,
            Designation=e.get("designation") or None,
            Worker_Type=e.get("worker_type") or None,
            Company=e.get("company") or None,
            status="active", created_by="mh-import")
        await session.execute(stmt.on_conflict_do_update(
            index_elements=["ID_Number"],
            set_={"Name": stmt.excluded.Name, "Site_ID": stmt.excluded.Site_ID,
                  "Designation": stmt.excluded.Designation,
                  "Worker_Type": stmt.excluded.Worker_Type,
                  "Company": stmt.excluded.Company,
                  "updated_at": _dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None)}))
        n += 1
    return n


@router.post("/import", summary="Import an attendance .xlsx (replace-by-date or append)")
async def import_attendance(file: UploadFile = File(...), replace: bool = True,
                            dry_run: bool = False, site_id: Optional[str] = None,
                            user: dict = Depends(require_roles("hod")),
                            session: AsyncSession = Depends(get_session)):
    sid = _write_site(user, site_id)
    data = await file.read()
    try:
        parsed = parse_attendance_workbook(data)
    except Exception as e:
        raise HTTPException(422, f"could not parse the workbook: {e}")
    if not parsed["employees"] and not parsed["timesheets"]:
        raise HTTPException(422, "no ADD EMPLOYEE / SAR rows found in the workbook")

    # Dates in the file that already hold rows for this site. Replace mode
    # deletes them first (predictable re-import); append mode would DUPLICATE
    # unassigned rows (NULL tags never conflict on the unique key), so the
    # overlap is surfaced here and the UI warns before importing.
    overlap: list[str] = []
    if parsed["dates"]:
        overlap = sorted({r[0] for r in (await session.execute(
            select(timesheets_t.c["Work_Date"]).distinct().where(
                timesheets_t.c["Site_ID"] == sid,
                timesheets_t.c["Work_Date"].in_(parsed["dates"]))))})

    if dry_run:
        return {"dry_run": True, "employees": len(parsed["employees"]),
                "timesheets": len(parsed["timesheets"]), "dates": parsed["dates"],
                "overlap_dates": overlap, "sample": parsed["timesheets"][:8]}

    if replace and parsed["dates"]:
        await session.execute(delete(timesheets_t).where(
            timesheets_t.c["Site_ID"] == sid,
            timesheets_t.c["Work_Date"].in_(parsed["dates"])))
    emp_n = 0
    for e in parsed["employees"]:
        stmt = pg_insert(employees_t).values(
            Site_ID=sid, Employee_Code=e["code"], Name=e["name"],
            Designation=e["designation"], Worker_Type=e["worker_type"],
            Company=e["company"], status="active", created_by="import",
            # QSEP: the join key back to the PERSON. This column has existed
            # since the baseline migration and was never written by anything
            # — which is exactly why a worker imported from the roster could
            # not be named on a supervisor material request.
            linked_id_number=e["code"])
        stmt = stmt.on_conflict_do_update(
            index_elements=["Site_ID", "Employee_Code"],
            set_={"Name": stmt.excluded.Name, "Designation": stmt.excluded.Designation,
                  "Worker_Type": stmt.excluded.Worker_Type,
                  "Company": stmt.excluded.Company,
                  "linked_id_number": stmt.excluded.linked_id_number,
                  "updated_at": _dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None)})
        await session.execute(stmt)
        emp_n += 1
    master_n = await _sync_employee_master(session, sid, parsed["employees"])
    ts_n = 0
    thr = await thresholds_for_codes(session, sid,
                                     [t["code"] for t in parsed["timesheets"]])
    gi = (await ot_thresholds(session))["GI"]
    for t in parsed["timesheets"]:
        await _upsert_timesheet(session, sid, t["code"], t["work_date"],
                                t["in_time"], t["out_time"], location=t["location"],
                                equipment_tag=t["equipment_tag"], system_code="",
                                status=t["status"], remarks=t["remarks"],
                                created_by="import",
                                threshold=thr.get(str(t["code"]).strip(), gi))
        ts_n += 1
    await write_audit(session, user["username"], "MH_IMPORT", "mh_timesheets",
                      f"{sid} employees={emp_n} master={master_n} timesheets={ts_n} "
                      f"replace={replace} dates={len(parsed['dates'])} "
                      f"overlap={len(overlap)}")
    await session.commit()
    return {"imported": True, "employees": emp_n, "employee_master": master_n,
            "timesheets": ts_n,
            "dates": parsed["dates"], "replace": replace,
            "overlap_dates": [] if replace else overlap}


# --- Exports (reuse the shared /reports renderers — DRY) -----------------------------
@router.get("/export/{key}", summary="Export a man-hours view (xlsx | csv | pdf)")
async def mh_export(key: str, format: str = "xlsx", site_id: Optional[str] = None,
                    employee_code: Optional[str] = None,
                    date_from: Optional[str] = None, date_to: Optional[str] = None,
                    user: dict = Depends(require_roles("hod")),
                    session: AsyncSession = Depends(get_session)):
    from fastapi.responses import StreamingResponse

    from .reports import _FORMATS
    fmt = format.lower()
    if fmt not in _FORMATS:
        raise HTTPException(400, f"format must be one of {sorted(_FORMATS)}")
    sid = resolve_site_param(user, site_id)

    if key == "employees":
        title = "MH Labor Roster"
        items = (await list_employees(site_id, None, user, session))["items"]
    elif key == "timesheets":
        title = "MH Timesheets"
        items = (await list_timesheets(site_id=site_id, employee_code=employee_code,
                                       date_from=date_from, date_to=date_to,
                                       user=user, session=session))["items"]
    elif key == "variance":
        title = "MH Estimate vs Actual"
        items = await _variance_rows(session, sid)
    elif key == "scorecard":
        title = "Equipment Scorecard (Material vs Labor)"
        items = (await _scorecard_rows(session, sid))["items"]
    elif key == "productivity":
        title = "MH Productivity Norms"
        items = (await _productivity_rows(session, sid))["items"]
    elif key == "employee-timeline":
        title = "MH Employee-wise Report"
        items = (await employee_timeline(site_id, employee_code, date_from,
                                         date_to, user, session))["items"]
    else:
        raise HTTPException(404, f"unknown man-hours export {key!r}")

    columns = list(items[0].keys()) if items else []
    rows = [[r.get(c) for c in columns] for r in items]
    render, media = _FORMATS[fmt]
    data = render(title, columns, rows, user["username"])
    return StreamingResponse(io.BytesIO(data), media_type=media,
                             headers={"Content-Disposition":
                                      f'attachment; filename="mh-{key}.{fmt}"'})
